from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, get_object_or_404
from django.views.generic import TemplateView, View
from django.core import serializers
from django.http import JsonResponse, HttpResponse, QueryDict
from django.db.models import Case, When, IntegerField, FloatField, ExpressionWrapper, Q, F, Sum, Count, IntegerField, Avg, Value, DecimalField
from django.db.models.functions import Cast, Round, Concat, Substr
from django.utils import timezone
from calendar import monthrange
import json
from datetime import date, datetime
from apps.main.models import Sector, Provincia, Distrito, Establecimiento, UPS, Profesion
from .models import padron_nom, sello, actas_homol, plano,cnv

# library excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color

import locale
import datetime
import os
import zipfile
from zipfile import ZIP_DEFLATED
import csv


# Create your views here.
class PadronView(TemplateView):
    template_name = 'pn/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['provincia'] = Provincia.objects.exclude(codigo__in=['00'])
        return context


class DistrictView(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']).order_by('nombre'), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListSello(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['seguro'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                avanxdist = sello.objects.filter(mide=1, medicion__in=list(eval(request.POST['mes']))).values('provincia', 'distrito').annotate(
                            den=Sum('den'), dni=Sum('var_dni'), direc=Sum('var_direc'))
                avanxdist = avanxdist.annotate(menor_valor=Case(When(dni__lt=F('direc'), then=F('dni')), When(dni__gte=F('direc'), then=F('direc')),
                            output_field=IntegerField())).order_by('provincia', 'distrito')

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                avanxdist = sello.objects.filter(cod_prov=request.POST['provincia'], mide=1, medicion__in=list(eval(request.POST['mes']))).values('provincia', 'distrito').annotate(
                            den=Sum('den'), dni=Sum('var_dni'), direc=Sum('var_direc'))
                avanxdist = avanxdist.annotate(menor_valor=Case(When(dni__lt=F('direc'), then=F('dni')), When(dni__gte=F('direc'), then=F('direc')),
                            output_field=IntegerField())).order_by('provincia', 'distrito')

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                avanxdist = sello.objects.filter(cod_dist=request.POST['distrito'], mide=1, medicion__in=list(eval(request.POST['mes']))).values('provincia', 'distrito').annotate(
                            den=Sum('den'), dni=Sum('var_dni'), direc=Sum('var_direc'))
                avanxdist = avanxdist.annotate(menor_valor=Case(When(dni__lt=F('direc'), then=F('dni')), When(dni__gte=F('direc'), then=F('direc')),
                            output_field=IntegerField())).order_by('provincia', 'distrito')

        else:
            if request.POST['provincia'] == 'TODOS':
                avanxdist = sello.objects.filter(seguro=request.POST['seguro'], mide=1, medicion__in=list(eval(request.POST['mes']))).values('provincia', 'distrito').annotate(
                            den=Sum('den'), dni=Sum('var_dni'), direc=Sum('var_direc'))
                avanxdist = avanxdist.annotate(menor_valor=Case(When(dni__lt=F('direc'), then=F('dni')), When(dni__gte=F('direc'), then=F('direc')),
                            output_field=IntegerField())).order_by('provincia', 'distrito')

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                avanxdist = sello.objects.filter(cod_prov=request.POST['provincia'], seguro=request.POST['seguro'], mide=1, medicion__in=list(eval(request.POST['mes']))).values('provincia', 'distrito').annotate(
                            den=Sum('den'), dni=Sum('var_dni'), direc=Sum('var_direc'))
                avanxdist = avanxdist.annotate(menor_valor=Case(When(dni__lt=F('direc'), then=F('dni')), When(dni__gte=F('direc'), then=F('direc')),
                            output_field=IntegerField())).order_by('provincia', 'distrito')

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                avanxdist = sello.objects.filter(cod_dist=request.POST['distrito'], seguro=request.POST['seguro'], mide=1, medicion__in=list(eval(request.POST['mes']))).values('provincia', 'distrito').annotate(
                            den=Sum('den'), dni=Sum('var_dni'), direc=Sum('var_direc'))
                avanxdist = avanxdist.annotate(menor_valor=Case(When(dni__lt=F('direc'), then=F('dni')), When(dni__gte=F('direc'), then=F('direc')),
                            output_field=IntegerField())).order_by('provincia', 'distrito')

        total = 0; menor = 0
        for avance in avanxdist:
            total = total + avance['den']
            menor = menor + avance['menor_valor']

        if total == 0:
            dataTotal = {'total': 0, 'cumple': 0, 'avance': 0}
        else:
            dataTotal = { 'total': total, 'cumple': menor, 'avance': round((menor / total)*100, 1)}

        json_data4.append(dataTotal)
        json_data4.append(list(avanxdist))
        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintSello(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:O2", "medium", "2F75B5")
        set_border(self, ws, "A3:O3", "medium", "366092")
        set_border(self, ws, "A5:O5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 23
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 31
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 6
        ws.column_dimensions['N'].width = 4
        ws.column_dimensions['O'].width = 8

        ws.merge_cells('B2:O2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: NOMINAL SELLO MUNICIPAL'

        ws.merge_cells('B3:O3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'Fuente: BD_PADRON_NOMINAL con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A5'] = '#'
        ws['A5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B5'] = 'Provincia'
        ws['B5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C5'] = 'Distrito'
        ws['C5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D5'] = 'Documento'
        ws['D5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E5'] = 'Fecha Nacido'
        ws['E5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F5'] = 'Menor Enc.'
        ws['F5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G5'] = 'CCPP'
        ws['G5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H5'] = 'Eje Vial'
        ws['H5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['I5'] = 'Descripción'
        ws['I5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['J5'] = 'Referencia'
        ws['J5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['K5'] = 'Seguro'
        ws['K5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['L5'] = 'Prog. Soc.'
        ws['L5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['M5'] = 'Estado'
        ws['M5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['N5'] = 'Dni'
        ws['N5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['N5'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['O5'] = 'Dirección'
        ws['O5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['O5'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        if request.GET['seguro'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = sello.objects.filter(mide=1, medicion__in=list(eval(request.GET['mes']))).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = sello.objects.filter(mide=1, medicion__in=list(eval(request.GET['mes'])), cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = sello.objects.filter(mide=1, medicion__in=list(eval(request.GET['mes'])), cod_dist=request.GET['dist']).order_by('provincia')
        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = sello.objects.filter(mide=1, medicion__in=list(eval(request.GET['mes'])), seguro=request.GET['seguro']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = sello.objects.filter(mide=1, medicion__in=list(eval(request.GET['mes'])), seguro=request.GET['seguro'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = sello.objects.filter(mide=1, medicion__in=list(eval(request.GET['mes'])), seguro=request.GET['seguro'], cod_dist=request.GET['dist']).order_by('provincia')

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 6
        cant = len(dataNom)
        num=1
        if cant > 0:
            for sel in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = sel['fields']['provincia']
                ws.cell(row=cont, column=3).value = sel['fields']['distrito']
                ws.cell(row=cont, column=4).value = sel['fields']['documento']
                ws.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=5).value = sel['fields']['fec_nac']
                ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=6).value = sel['fields']['menor_encont']
                ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=7).value = sel['fields']['area_ccpp']
                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).value = sel['fields']['eje_vial']
                ws.cell(row=cont, column=8).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=9).value = sel['fields']['descripcion']
                ws.cell(row=cont, column=9).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=10).value = sel['fields']['ref_direc']
                ws.cell(row=cont, column=10).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=11).value = sel['fields']['seguro']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).value = sel['fields']['tprog_social']
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).value = sel['fields']['est_regist']
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                if sel['fields']['var_dni'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=14).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=14).font = Font(name='Calibri', size=10, color='C00000')
                ws.cell(row=cont, column=14).value = cumplen
                ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")

                if sel['fields']['var_direc'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=15).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=15).font = Font(name='Calibri', size=10, color='C00000')
                ws.cell(row=cont, column=15).value = cumplen
                ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO NOMINAL SELLO MUNICIPAL.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL SELLO'
        wb.save(response)
        return response


class ActasView(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', actas_homol.objects.all(), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class PrintPadronNom(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:AF2", "medium", "2F75B5")
        set_border(self, ws, "A3:AF3", "medium", "366092")
        set_border(self, ws, "A5:AF5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 23
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 11
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 13
        ws.column_dimensions['N'].width = 7
        ws.column_dimensions['O'].width = 11
        ws.column_dimensions['P'].width = 7
        ws.column_dimensions['Q'].width = 35
        ws.column_dimensions['R'].width = 35
        ws.column_dimensions['S'].width = 12
        ws.column_dimensions['T'].width = 32
        ws.column_dimensions['U'].width = 35
        ws.column_dimensions['V'].width = 35
        ws.column_dimensions['X'].width = 12
        ws.column_dimensions['Y'].width = 18
        ws.column_dimensions['Z'].width = 12
        ws.column_dimensions['AA'].width = 26
        ws.column_dimensions['AB'].width = 12
        ws.column_dimensions['AC'].width = 18
        ws.column_dimensions['AD'].width = 11
        ws.column_dimensions['AE'].width = 19
        ws.column_dimensions['AF'].width = 7

        ws.merge_cells('B2:AF2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: Padron Nominal'

        ws.merge_cells('B3:AF3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A5'] = '#'
        ws['A5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B5'] = 'Cod. Padron'
        ws['B5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C5'] = 'Provincia'
        ws['C5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D5'] = 'Distrito'
        ws['D5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E5'] = 'Establecimiento'
        ws['E5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F5'] = 'Centro Poblado'
        ws['F5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G5'] = 'Est. Trámite DNI'
        ws['G5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H5'] = 'Fecha Trámite'
        ws['H5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['I5'] = 'Tipo Doc'
        ws['I5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['J5'] = 'Documento'
        ws['J5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['K5'] = 'Apellidos y Nombres'
        ws['K5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['L5'] = 'Fecha Nacido'
        ws['L5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['M5'] = 'Menor Visit.'
        ws['M5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['N5'] = 'Menor Enc.'
        ws['N5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['N5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['O5'] = 'Seguro'
        ws['O5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['O5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['P5'] = 'Prog. Soc.'
        ws['P5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['P5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['Q5'] = 'Eje Vial'
        ws['Q5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Q5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['R5'] = 'Descripción'
        ws['R5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['R5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['R5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['S5'] = 'Fecha Visita'
        ws['S5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['S5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['S5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['T5'] = 'Fuente Datos'
        ws['T5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['T5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['T5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['U5'] = 'EESS Nacido'
        ws['U5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['U5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['U5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['V5'] = 'EESS Adscripción'
        ws['V5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['V5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['V5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['V5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['W5'] = 'Institución'
        ws['W5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['W5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['W5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['X5'] = 'Dni Madre'
        ws['X5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['X5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['X5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['X5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['Y5'] = 'Ape. Nom. Madre'
        ws['Y5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Y5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Y5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Y5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['Z5'] = 'Cel. Madre'
        ws['Z5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Z5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Z5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Z5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['AA5'] = 'Grado Inst.'
        ws['AA5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AA5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AA5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AA5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['AB5'] = 'Dni Jefe'
        ws['AB5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AB5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AB5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AB5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['AC5'] = 'Ape. Nom. Jefe'
        ws['AC5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AC5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AC5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AC5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['AD5'] = 'Entidad'
        ws['AD5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AD5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AD5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AD5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['AE5'] = 'Tipo Regist.'
        ws['AE5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AE5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AE5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AE5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['AF5'] = 'Est. Regist.'
        ws['AF5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AF5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AF5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['AF5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['prov'] == 'TODOS':
            dataNom = padron_nom.objects.all().order_by('provincia')
        elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
            dataNom = padron_nom.objects.filter(cod_prov=request.GET['prov']).order_by('provincia')
        elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
            dataNom = padron_nom.objects.filter(cod_dist=request.GET['dist']).order_by('provincia')
        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 6
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=2).value = pn['fields']['cod_padron']
                ws.cell(row=cont, column=3).value = pn['fields']['provincia']
                ws.cell(row=cont, column=4).value = pn['fields']['distrito']
                ws.cell(row=cont, column=5).value = pn['fields']['eess']
                ws.cell(row=cont, column=6).value = pn['fields']['ccpp']
                ws.cell(row=cont, column=7).value = pn['fields']['est_tramite']
                ws.cell(row=cont, column=8).value = pn['fields']['fec_tramite']
                ws.cell(row=cont, column=9).value = pn['fields']['tipo_doc']
                ws.cell(row=cont, column=10).value = pn['fields']['cnv_dni']
                ws.cell(row=cont, column=11).value = pn['fields']['nombres_ninio']
                ws.cell(row=cont, column=12).value = pn['fields']['fec_nac']
                ws.cell(row=cont, column=13).value = pn['fields']['menor_visit']
                ws.cell(row=cont, column=14).value = pn['fields']['menor_encont']
                ws.cell(row=cont, column=15).value = pn['fields']['seguro']
                ws.cell(row=cont, column=16).value = pn['fields']['tprog_social']
                ws.cell(row=cont, column=17).value = pn['fields']['eje_vial']
                ws.cell(row=cont, column=18).value = pn['fields']['descripcion']
                ws.cell(row=cont, column=19).value = pn['fields']['fec_visita']
                ws.cell(row=cont, column=20).value = pn['fields']['fuente']
                ws.cell(row=cont, column=21).value = pn['fields']['eess_nacido']
                ws.cell(row=cont, column=22).value = pn['fields']['eess_adscrip']
                ws.cell(row=cont, column=23).value = pn['fields']['institucion']
                ws.cell(row=cont, column=24).value = pn['fields']['dni_madre']
                ws.cell(row=cont, column=25).value = pn['fields']['nombres_madre']
                ws.cell(row=cont, column=26).value = pn['fields']['celular_madre']
                ws.cell(row=cont, column=27).value = pn['fields']['grado_inst']
                ws.cell(row=cont, column=28).value = pn['fields']['dni_jefe']
                ws.cell(row=cont, column=29).value = pn['fields']['nombres_jefe']
                ws.cell(row=cont, column=30).value = pn['fields']['entidad']
                ws.cell(row=cont, column=31).value = pn['fields']['tregistro']
                ws.cell(row=cont, column=32).value = pn['fields']['est_regist']

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO PADRON NOMINAL.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'PADRON NOMINAL'
        wb.save(response)
        return response


class PlanoView(TemplateView):
    template_name = 'plane/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['provincia'] = Provincia.objects.exclude(codigo__in=['00'])
        context['ups'] = UPS.objects.all().order_by('nombre')
        context['profesion'] = Profesion.objects.all().order_by('nombre')
        return context


class EESS(View):
    def get(self, request, *args, **kwargs):
        eess = serializers.serialize('json', Establecimiento.objects.filter(dist_id=request.GET['id'], sector_id=7).order_by('nombre'), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(eess, content_type='application/json')


def compress_file(file_data, filename):
    zip_filename = f"{filename}.zip"
    with zipfile.ZipFile(zip_filename, 'w', ZIP_DEFLATED) as zipf:
        zipf.writestr(filename, file_data.getvalue())
    return zip_filename


class PrintPlano(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        if request.GET['ups'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = plano.objects.filter(fec_aten__year=request.GET['anio'], mes=request.GET['mes']).annotate(doc_personal2=Concat(Substr('doc_personal', 1, 6), Value('** '))).order_by('-dia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = plano.objects.filter(cod_prov=request.GET['prov'], fec_aten__year=request.GET['anio'], mes=request.GET['mes']).annotate(doc_personal2=Concat(Substr('doc_personal', 1, 6), Value('** '))).order_by('-dia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = plano.objects.filter(cod_dist=request.GET['dist'], fec_aten__year=request.GET['anio'], mes=request.GET['mes']).annotate(doc_personal2=Concat(Substr('doc_personal', 1, 6), Value('** '))).order_by('-dia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = plano.objects.filter(cod_eess=request.GET['eess'], fec_aten__year=request.GET['anio'], mes=request.GET['mes']).annotate(doc_personal2=Concat(Substr('doc_personal', 1, 6), Value('** ')))

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = plano.objects.filter(id_ups=request.GET['ups'], fec_aten__year=request.GET['anio'], mes=request.GET['mes']).annotate(doc_personal2=Concat(Substr('doc_personal', 1, 6), Value('** '))).order_by('-dia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = plano.objects.filter(id_ups=request.GET['ups'], cod_prov=request.GET['prov'], fec_aten__year=request.GET['anio'], mes=request.GET['mes']).annotate(doc_personal2=Concat(Substr('doc_personal', 1, 6), Value('** '))).order_by('-dia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = plano.objects.filter(id_ups=request.GET['ups'], cod_dist=request.GET['dist'], fec_aten__year=request.GET['anio'], mes=request.GET['mes']).annotate(doc_personal2=Concat(Substr('doc_personal', 1, 6), Value('** '))).order_by('-dia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = plano.objects.filter(id_ups=request.GET['ups'], cod_eess=request.GET['eess'], fec_aten__year=request.GET['anio'], mes=request.GET['mes']).annotate(doc_personal2=Concat(Substr('doc_personal', 1, 6), Value('** '))).order_by('-dia')


        nombre_archivo = "archivo_plano.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        writer = csv.writer(response)
        writer.writerow(['Id Cita', 'Lote', 'Mes', 'Dia', 'Fec Atencion','Num Pag','Desc UPS','Sector','Red','Provincia','MicroRed',
                         'Distrito','Cod Unico','EESS','Tipo Doc','Doc Paciente','Nombres Paciente','Fec Nac Paciente','Etnia',
                         'Genero','Hist Clínica','Ficha Familiar','Financiador','Pais','Doc Personal','Profesion', 'Doc Registrador',
                         'id cond eess','Edad Reg','Tipo Edad','Grupo Edad','Codigo','Tipo Diag', 'Desc Item','Lab','Id Corr Lab',
                         'Peso','Talla','Hemoglobina', 'Desc Otra Cond','Desc CCPP','FUR','Fec Solic HB','Fec Result HB','Fecha Registro',
                         'Fec Modificacion'])

        for item in dataNom:
            writer.writerow([item.id_cita, item.lote, item.mes, item.dia, item.fec_aten, item.num_pag, item.desc_ups, item.desc_sector,
                             item.red, item.provincia, item.microred, item.distrito, item.cod_unico, item.eess, item.tdoc_pacien,
                             item.doc_pacien, item.nombres_pacien, item.fnac_pacien, item.desc_etnia, item.genero, item.his_clinica,
                             item.ficha_fam, item.financiador, item.pais, item.doc_personal2, item.profesion, item.doc_regist,
                             item.id_cond_eess, item.edad_reg, item.tedad, item.grupo_edad, item.codigo, item.tdiag, item.desc_item,
                             item.vlab, item.id_corr_lab, item.peso, item.talla, item.hb, item.dec_otra_cond, item.dec_ccpp,
                             item.fur, item.solic_hb, item.result_hb, item.fregistro, item.fmodific])

        zip_filename = compress_file(response, nombre_archivo)

        with open(zip_filename, 'rb') as zip_file:
            response = HttpResponse(zip_file.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            return response


class R40View(TemplateView):
    template_name = 'r40/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['provincia'] = Provincia.objects.exclude(codigo__in=['00'])
        context['profesion'] = UPS.objects.all().order_by('nombre')
        return context


class PrintR40Prof(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:O2", "medium", "2F75B5")
        set_border(self, ws, "A3:O3", "medium", "366092")
        set_border(self, ws, "A5:O5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 23
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 9
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 9
        ws.column_dimensions['O'].width = 9

        ws.merge_cells('B2:O2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: R40 POR PROFESIÓN - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:O3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A5'] = '#'
        ws['A5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B5'] = 'Provincia'
        ws['B5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C5'] = 'Distrito'
        ws['C5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D5'] = 'Establecimiento'
        ws['D5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E5'] = 'Documento'
        ws['E5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F5'] = 'UPS'
        ws['F5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G5'] = 'Fecha atención'
        ws['G5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H5'] = 'Atendidos'
        ws['H5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H5'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['I5'] = 'Aten Serv'
        ws['I5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I5'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['J5'] = 'Atenciones'
        ws['J5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J5'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['K5'] = 'Nuevo'
        ws['K5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K5'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['L5'] = 'Conti'
        ws['L5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L5'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['M5'] = 'Reing'
        ws['M5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M5'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['N5'] = 'APP'
        ws['N5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['N5'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['O5'] = 'AAA'
        ws['O5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['O5'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        if request.GET['prov'] == 'TODOS':
            result = (plano.objects.filter(id_corr_item=1, id_corr_lab=1, anio=request.GET['anio'], mes=request.GET['mes'], id_prof=request.GET['prof'])
                    .values( 'provincia', 'distrito', 'eess', 'doc_personal', 'profesion', 'desc_ups', 'fec_aten')
                    .annotate(
                        ATEND=Sum(Case(When(id_pac__isnull=False, id_cond_eess__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                        ATENC=Sum(Case(When(id_pac__isnull=False, id_cond_serv__in=['N', 'R', 'C'], then=1), default=0, output_field=IntegerField())),
                        ATEND_SER_TOTAL=Sum(Case(When(id_pac__isnull=False, id_cond_serv__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                        NUEVO=Sum(Case(When(id_pac__isnull=False, id_cond_serv='N', then=1), default=0, output_field=IntegerField())),
                        CONTI=Sum(Case(When(id_pac__isnull=False, id_cond_serv='C', then=1), default=0, output_field=IntegerField())),
                        REING=Sum(Case(When(id_pac__isnull=False, id_cond_serv='R', then=1), default=0, output_field=IntegerField())),
                        APP=Sum(Case(When(id_pac__startswith='APP', then=1), default=0, output_field=IntegerField())),
                        AAA=Sum(Case(When(id_pac__startswith='AAA', then=1), default=0, output_field=IntegerField()))
                    ).order_by('provincia', 'distrito', 'eess'))

        elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
            result = (plano.objects.filter(id_corr_item=1, id_corr_lab=1, anio=request.GET['anio'], mes=request.GET['mes'], id_prof=request.GET['prof'], cod_prov=request.GET['prov'])
                    .values( 'provincia', 'distrito', 'eess', 'doc_personal', 'profesion', 'desc_ups', 'fec_aten')
                    .annotate(
                        ATEND=Sum(Case(When(id_pac__isnull=False, id_cond_eess__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                        ATENC=Sum(Case(When(id_pac__isnull=False, id_cond_serv__in=['N', 'R', 'C'], then=1), default=0, output_field=IntegerField())),
                        ATEND_SER_TOTAL=Sum(Case(When(id_pac__isnull=False, id_cond_serv__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                        NUEVO=Sum(Case(When(id_pac__isnull=False, id_cond_serv='N', then=1), default=0, output_field=IntegerField())),
                        CONTI=Sum(Case(When(id_pac__isnull=False, id_cond_serv='C', then=1), default=0, output_field=IntegerField())),
                        REING=Sum(Case(When(id_pac__isnull=False, id_cond_serv='R', then=1), default=0, output_field=IntegerField())),
                        APP=Sum(Case(When(id_pac__startswith='APP', then=1), default=0, output_field=IntegerField())),
                        AAA=Sum(Case(When(id_pac__startswith='AAA', then=1), default=0, output_field=IntegerField()))
                    ).order_by('provincia', 'distrito', 'eess'))

        elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
            result = (plano.objects.filter(id_corr_item=1, id_corr_lab=1, anio=request.GET['anio'], mes=request.GET['mes'], id_prof=request.GET['prof'], cod_dist=request.GET['dist'])
                    .values( 'provincia', 'distrito', 'eess', 'doc_personal', 'profesion', 'desc_ups', 'fec_aten')
                    .annotate(
                        ATEND=Sum(Case(When(id_pac__isnull=False, id_cond_eess__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                        ATENC=Sum(Case(When(id_pac__isnull=False, id_cond_serv__in=['N', 'R', 'C'], then=1), default=0, output_field=IntegerField())),
                        ATEND_SER_TOTAL=Sum(Case(When(id_pac__isnull=False, id_cond_serv__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                        NUEVO=Sum(Case(When(id_pac__isnull=False, id_cond_serv='N', then=1), default=0, output_field=IntegerField())),
                        CONTI=Sum(Case(When(id_pac__isnull=False, id_cond_serv='C', then=1), default=0, output_field=IntegerField())),
                        REING=Sum(Case(When(id_pac__isnull=False, id_cond_serv='R', then=1), default=0, output_field=IntegerField())),
                        APP=Sum(Case(When(id_pac__startswith='APP', then=1), default=0, output_field=IntegerField())),
                        AAA=Sum(Case(When(id_pac__startswith='AAA', then=1), default=0, output_field=IntegerField()))
                    ).order_by('provincia', 'distrito', 'eess'))

        elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
            result = (plano.objects.filter(id_corr_item=1, id_corr_lab=1, anio=request.GET['anio'], mes=request.GET['mes'], id_prof=request.GET['prof'], cod_eess=request.GET['eess'])
                    .values( 'provincia', 'distrito', 'eess', 'doc_personal', 'profesion', 'desc_ups', 'fec_aten')
                    .annotate(
                        ATEND=Sum(Case(When(id_pac__isnull=False, id_cond_eess__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                        ATENC=Sum(Case(When(id_pac__isnull=False, id_cond_serv__in=['N', 'R', 'C'], then=1), default=0, output_field=IntegerField())),
                        ATEND_SER_TOTAL=Sum(Case(When(id_pac__isnull=False, id_cond_serv__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                        NUEVO=Sum(Case(When(id_pac__isnull=False, id_cond_serv='N', then=1), default=0, output_field=IntegerField())),
                        CONTI=Sum(Case(When(id_pac__isnull=False, id_cond_serv='C', then=1), default=0, output_field=IntegerField())),
                        REING=Sum(Case(When(id_pac__isnull=False, id_cond_serv='R', then=1), default=0, output_field=IntegerField())),
                        APP=Sum(Case(When(id_pac__startswith='APP', then=1), default=0, output_field=IntegerField())),
                        AAA=Sum(Case(When(id_pac__startswith='AAA', then=1), default=0, output_field=IntegerField()))
                    ).order_by('provincia', 'distrito', 'eess'))


        cont = 6
        cant = len(list(result))
        num=1
        if cant > 0:
            for rprof in list(result):
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = rprof['provincia']
                ws.cell(row=cont, column=3).value = rprof['distrito']
                ws.cell(row=cont, column=4).value = rprof['eess']
                ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=5).value = rprof['doc_personal']
                ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=6).value = rprof['desc_ups']
                ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=7).value = rprof['fec_aten']
                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).value = rprof['ATEND']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = rprof['ATEND_SER_TOTAL']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = rprof['ATENC']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).value = rprof['NUEVO']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).value = rprof['CONTI']
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).value = rprof['REING']
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=14).value = rprof['APP']
                ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=15).value = rprof['AAA']
                ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO R40 X PROFESIONAL.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL R40'
        wb.save(response)
        return response


class PrintR40Doc(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:H2", "medium", "2F75B5")
        set_border(self, ws, "A3:H3", "medium", "366092")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 23
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 6
        ws.column_dimensions['F'].width = 6
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 11

        ws.merge_cells('B2:H2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: ATENDIDOS - ATENCIONES'

        ws.merge_cells('B3:H3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A5'] = '#'
        ws['A5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B5'] = 'Provincia'
        ws['B5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C5'] = 'Distrito'
        ws['C5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D5'] = 'Establecimiento'
        ws['D5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E5'] = 'Año'
        ws['E5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F5'] = 'Mes'
        ws['F5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G5'] = 'Atendidos'
        ws['G5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H5'] = 'Atenciones'
        ws['H5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H5'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

        start_date = timezone.datetime(int(request.GET['anioIni']), int(request.GET['mesIni']), 1)
        end_get = timezone.datetime(int(request.GET['anioFin']), int(request.GET['mesFin']), 1)
        end_date = timezone.datetime(int(request.GET['anioFin']), int(request.GET['mesFin']), monthrange(end_get.year, end_get.month)[1])

        result = (plano.objects.filter(id_corr_item=1, id_corr_lab=1, id_pac__isnull=False, doc_personal=request.GET['doc'], fec_aten__range=(start_date, end_date))
                .values('provincia', 'distrito', 'eess', 'anio', 'mes')
                .annotate(
                    Atend_eess=Sum(Case(When(id_cond_eess__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                    Atend_Serv=Sum(Case(When(id_cond_serv__in=['N', 'R'], then=1), default=0, output_field=IntegerField())),
                    Atenciones=Sum(Case(When(id_cond_serv__in=['N', 'R', 'C'], then=1), default=0, output_field=IntegerField()))
                ))

        cont = 6
        cant = len(result)
        num=1
        if cant > 0:
            for atendidos in list(result):
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = atendidos['provincia']
                ws.cell(row=cont, column=3).value = atendidos['distrito']
                ws.cell(row=cont, column=4).value = atendidos['eess']
                ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=5).value = atendidos['anio']
                ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=6).value = atendidos['mes']
                ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=7).value = atendidos['Atend_Serv']
                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).value = atendidos['Atenciones']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO ATENDIDOS ATENCIONES.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL ATENDIDOS ATENCIONES'
        wb.save(response)
        return response


class PrintCnv(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:AB2", "medium", "2F75B5")
        set_border(self, ws, "A3:AB3", "medium", "366092")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 23
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['M'].width = 6
        ws.column_dimensions['N'].width = 14
        ws.column_dimensions['O'].width = 11
        ws.column_dimensions['R'].width = 26
        ws.column_dimensions['S'].width = 29
        ws.column_dimensions['T'].width = 27
        ws.column_dimensions['U'].width = 13
        ws.column_dimensions['X'].width = 15
        ws.column_dimensions['Y'].width = 24
        ws.column_dimensions['Z'].width = 31
        ws.column_dimensions['AA'].width = 25

        ws.merge_cells('B2:AB2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO DEIT: NOMINAL CNV (Se considera todos los nacimientos dentro y fuera de la región asignados a Pasco)'

        ws.merge_cells('B3:AB3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'Fuente: BD_CNV con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A5'] = '#'
        ws['A5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B5'] = 'Provincia'
        ws['B5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C5'] = 'Distrito'
        ws['C5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D5'] = 'Cod EESS'
        ws['D5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E5'] = 'Establecimiento'
        ws['E5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E5'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F5'] = 'Documento'
        ws['F5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G5'] = 'Fecha Nacido'
        ws['G5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H5'] = 'Institución'
        ws['H5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['I5'] = 'Categoria'
        ws['I5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['J5'] = 'Periodo'
        ws['J5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['K5'] = 'Peso'
        ws['K5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['L5'] = 'Talla'
        ws['L5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['M5'] = 'Sem Gest'
        ws['M5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['N5'] = 'Condición Parto'
        ws['N5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['O5'] = 'Sexo'
        ws['O5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['P5'] = 'Tipo Parto'
        ws['P5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['Q5'] = 'Financiador'
        ws['Q5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['R5'] = 'Prof Certif'
        ws['R5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['R5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['S5'] = 'Certifica Que'
        ws['S5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['S5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['T5'] = 'Lugar Nacido'
        ws['T5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['T5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['U5'] = 'Estado Civil'
        ws['U5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['U5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['V5'] = 'TDoc Madre'
        ws['V5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['V5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['V5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['V5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['W5'] = 'Doc Madre'
        ws['W5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['W5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['W5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['X5'] = 'Dpto Madre'
        ws['X5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['X5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['X5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['X5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['Y5'] = 'Provincia Madre'
        ws['Y5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Y5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Y5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Y5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['Z5'] = 'Distrito Madre'
        ws['Z5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Z5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Z5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Z5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['AA5'] = 'Fecha Reg'
        ws['AA5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AA5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AA5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AA5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['AB5'] = 'Apgar'
        ws['AB5'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AB5'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AB5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AB5'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if len(request.GET['mes']) == 1:
            mes = '0'+request.GET['mes']
        else:
            mes = request.GET['mes']

        if request.GET['mes'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = cnv.objects.filter(cod_dep=27, periodo__startswith=request.GET['anio'])
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = cnv.objects.filter(cod_dep=27, cod_prov=request.GET['prov'], periodo__startswith=request.GET['anio'])
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = cnv.objects.filter(cod_dep=27, cod_dist=request.GET['dist'], periodo__startswith=request.GET['anio'])

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = cnv.objects.filter(cod_dep=27, periodo=str(request.GET['anio']+mes))
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = cnv.objects.filter(cod_dep=27, periodo=str(request.GET['anio']+mes), cod_prov=request.GET['prov'])
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = cnv.objects.filter(cod_dep=27, periodo=str(request.GET['anio']+mes), cod_dist=request.GET['dist'])

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 6
        cant = len(dataNom)
        num=1
        if cant > 0:
            for nomcnv in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = nomcnv['fields']['provincia']
                ws.cell(row=cont, column=3).value = nomcnv['fields']['distrito']
                ws.cell(row=cont, column=4).value = nomcnv['fields']['cod_eess']
                ws.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=5).value = nomcnv['fields']['eess']
                ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=6).value = nomcnv['fields']['cnv']
                ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=7).value = nomcnv['fields']['fec_nac']
                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).value = nomcnv['fields']['institucion']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = nomcnv['fields']['categoria']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = nomcnv['fields']['periodo']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).value = nomcnv['fields']['peso']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).value = nomcnv['fields']['talla']
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).value = nomcnv['fields']['dur_emb']
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=14).value = nomcnv['fields']['cond_parto']
                ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=15).value = nomcnv['fields']['sexo']
                ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=16).value = nomcnv['fields']['tparto']
                ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=17).value = nomcnv['fields']['financiador']
                ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=18).value = nomcnv['fields']['prof_cert']
                ws.cell(row=cont, column=19).value = nomcnv['fields']['cert_prof']
                ws.cell(row=cont, column=20).value = nomcnv['fields']['lugar_nac']
                ws.cell(row=cont, column=21).value = nomcnv['fields']['est_civil']
                ws.cell(row=cont, column=21).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=22).value = nomcnv['fields']['tdoc_madre']
                ws.cell(row=cont, column=22).alignment = Alignment(horizontal="center", wrap_text=True)
                ws.cell(row=cont, column=23).value = nomcnv['fields']['doc_madre']
                ws.cell(row=cont, column=23).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=24).value = nomcnv['fields']['dpt_madre']
                ws.cell(row=cont, column=25).value = nomcnv['fields']['prov_madre']
                ws.cell(row=cont, column=26).value = nomcnv['fields']['dist_madre']
                ws.cell(row=cont, column=27).value = nomcnv['fields']['fe_crea']
                ws.cell(row=cont, column=27).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=28).value = nomcnv['fields']['apgar']
                ws.cell(row=cont, column=28).alignment = Alignment(horizontal="center")

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO NOMINAL CNV.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL CNV'
        wb.save(response)
        return response

