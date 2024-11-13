from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, get_object_or_404
from django.views.generic import TemplateView, View
from django.core import serializers
from django.http import JsonResponse, HttpResponse, QueryDict
from django.db.models import Case, When, IntegerField, FloatField, ExpressionWrapper, Q, F, Sum, Count, IntegerField, Avg, Value, DecimalField
from django.db.models.functions import Cast, Round
import json
from datetime import date, datetime

from .models import mc_03, si_01, si_0202, si_0203, si_0401, vii_0101, vi_0101, vi_0102, si_0201_pr, si_0201_sn, si_0201_cont
from apps.main.models import Sector, Provincia, Distrito, Establecimiento

# library excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color

import locale
import datetime

class MC01View(TemplateView):
    template_name = 'pregnant/MC-01.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['provincia'] = Provincia.objects.all()
        return context


class MC02View(TemplateView):
    template_name = 'child/MC-02.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['provincia'] = Provincia.objects.all()
        return context


class MC03View(TemplateView):
    template_name = 'child/MC-03.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['provincia'] = Provincia.objects.all()
        return context


class DistrictView(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListMC03(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['provincia'] == 'TODOS':
            total = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
            cumplen = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
            dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
            dataProv = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataDist = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataNom = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
            total = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
            cumplen = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
            dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
            dataProv = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataDist = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataNom = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
            print(request.POST['distrito'])
            total = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
            cumplen = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
            dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
            dataProv = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataDist = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                        numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                        output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            dataNom = mc_03.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintMC03(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:Q2", "medium", "2F75B5")
        set_border(self, ws, "A3:Q3", "medium", "366092")
        set_border(self, ws, "A5:Q5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 23
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['Q'].width = 18

        ws.merge_cells('B2:Q2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: MC-03: Recién nacidos del departamento, reciben vacunas BCG, HvB, controles CRED y tamizaje neonatal - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:Q3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'CODIFICACION: Cred Rn: 99381.01 - Tamizaje: 36416 - BCG: 90585 - HVB: 90744'

        ws.merge_cells('A5:Q5')
        ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Provincia'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Distrito'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Ult. EESS His'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Documento'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Fecha Nacido'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Seguro'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Tmz'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

        ws['I7'] = 'Bcg'
        ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['J7'] = 'Hvb'
        ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['K7'] = 'Ctrl 1'
        ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K7'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['L7'] = 'Visita 7d'
        ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['M7'] = 'Ctrl 2'
        ws['M7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M7'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['N7'] = 'Ctrl 3'
        ws['N7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['N7'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['O7'] = 'Ctrl 4'
        ws['O7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['O7'].fill = PatternFill(start_color='E2F5CB', end_color='E2F5CB', fill_type='solid')

        ws['P7'] = 'Cumple'
        ws['P7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['P7'].fill = PatternFill(start_color='b3f5c2', end_color='b3f5c2', fill_type='solid')

        ws['Q7'] = 'Programa'
        ws['Q7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Q7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['prov'] == 'TODOS':
            dataNom = mc_03.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
        elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
            dataNom = mc_03.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
        elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
            dataNom = mc_03.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('provincia')
        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 8
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pqtrn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = pqtrn['fields']['provincia']
                ws.cell(row=cont, column=3).value = pqtrn['fields']['distrito']
                ws.cell(row=cont, column=4).value = pqtrn['fields']['ult_eess']
                ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=5).value = pqtrn['fields']['documento']
                ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=6).value = pqtrn['fields']['fec_nac']
                ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=7).value = pqtrn['fields']['seguro']
                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).value = pqtrn['fields']['tmz']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = pqtrn['fields']['bcg']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = pqtrn['fields']['hvb']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).value = pqtrn['fields']['ctrl1']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).value = pqtrn['fields']['visit7d_prom']
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).value = pqtrn['fields']['ctrl2']
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=14).value = pqtrn['fields']['ctrl3']
                ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=15).value = pqtrn['fields']['ctrl4']
                ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                if pqtrn['fields']['num'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=16).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=16).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=16).value = cumplen
                ws.cell(row=cont, column=17).value = pqtrn['fields']['programa']
                ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO PAQUETE RECIEN NACIDO.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL PAQUETE RN'
        wb.save(response)
        return response


class SI01View(TemplateView):
    template_name = 'pregnant/SI-01.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sector'] = Sector.objects.all()
        context['provincia'] = Provincia.objects.all()
        return context


class DistrictSIView(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListSI01(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['sector'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                total = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
                cumplen = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
                cumplen = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
                cumplen = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['provincia'] == 'TODOS':
                total = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_01.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))


        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintSI01(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:M2", "medium", "2F75B5")
        set_border(self, ws, "A3:M3", "medium", "366092")
        set_border(self, ws, "A5:M5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 35
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 8
        ws.column_dimensions['M'].width = 18

        ws.merge_cells('B2:M2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: SI-01: Gestantes con diagnóstico de anemia atendidas en eess de salud del I y II nivel de atención (con población asignada) del Gobierno Regional, que reciben dosaje de hemoglobina de control y segunda entrega de tto con hierro - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:M3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'CODIFICACION: Dosaje Hemoglobina: 85018, 85018.01 - Suplementación: 99199.26, 59401.04 - Dx Anemmia: O990'

        ws.merge_cells('A5:M5')
        ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Sector'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Provincia'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Distrito'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Establecimiento'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Ult. EESS His'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Documento'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Fecha Dx'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['I7'] = 'Suple 1'
        ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I7'].fill = PatternFill(start_color='DCCBF5', end_color='DCCBF5', fill_type='solid')

        ws['J7'] = 'Dosaje'
        ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

        ws['K7'] = 'Suple 2'
        ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K7'].fill = PatternFill(start_color='F5D4CB', end_color='F5D4CB', fill_type='solid')

        ws['L7'] = 'Cumple'
        ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['L7'].fill = PatternFill(start_color='b3f5c2', end_color='b3f5c2', fill_type='solid')

        ws['M7'] = 'Programa'
        ws['M7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['M7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['sector'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = si_01.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = si_01.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = si_01.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = si_01.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = si_01.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = si_01.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 8
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pqtrn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                ws.cell(row=cont, column=5).value = pqtrn['fields']['eess']
                ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=6).value = pqtrn['fields']['ult_eess']
                ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=7).value = pqtrn['fields']['documento']
                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).value = pqtrn['fields']['fecha_dx']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = pqtrn['fields']['suple1']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = pqtrn['fields']['dosaje']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).value = pqtrn['fields']['suple2']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                if pqtrn['fields']['num'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=12).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=12).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=12).value = cumplen
                ws.cell(row=cont, column=13).value = pqtrn['fields']['programa']
                ws.cell(row=cont, column=13).alignment = Alignment(wrap_text=True)

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO GESTANTES CON ANEMIA Y TTO.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL GESTANTES CON ANEMIA Y TTO'
        wb.save(response)
        return response


class SI0201View(TemplateView):
    template_name = 'child/SI-02_01.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sector'] = Sector.objects.all()
        context['provincia'] = Provincia.objects.all()
        return context


class ListSI0201(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['sector'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                total = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
                cumplen = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataPr = si_0201_pr.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataPr = json.loads(serializers.serialize('json', dataPr, indent=2, use_natural_foreign_keys=True))
                dataSn = si_0201_sn.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataSn = json.loads(serializers.serialize('json', dataSn, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
                cumplen = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataPr = si_0201_pr.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataPr = json.loads(serializers.serialize('json', dataPr, indent=2, use_natural_foreign_keys=True))
                dataSn = si_0201_sn.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataSn = json.loads(serializers.serialize('json', dataSn, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
                cumplen = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataPr = si_0201_pr.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataPr = json.loads(serializers.serialize('json', dataPr, indent=2, use_natural_foreign_keys=True))
                dataSn = si_0201_sn.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataSn = json.loads(serializers.serialize('json', dataSn, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['provincia'] == 'TODOS':
                total = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataPr = si_0201_pr.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataPr = json.loads(serializers.serialize('json', dataPr, indent=2, use_natural_foreign_keys=True))
                dataSn = si_0201_sn.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataSn = json.loads(serializers.serialize('json', dataSn, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataPr = si_0201_pr.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataPr = json.loads(serializers.serialize('json', dataPr, indent=2, use_natural_foreign_keys=True))
                dataSn = si_0201_sn.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataSn = json.loads(serializers.serialize('json', dataSn, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0201_cont.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataPr = si_0201_pr.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataPr = json.loads(serializers.serialize('json', dataPr, indent=2, use_natural_foreign_keys=True))
                dataSn = si_0201_sn.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataSn = json.loads(serializers.serialize('json', dataSn, indent=2, use_natural_foreign_keys=True))


        json_data4.append(dataPr)
        json_data4.append(dataSn)
        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintSI0201(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:V2", "medium", "2F75B5")
        set_border(self, ws, "A3:V3", "medium", "366092")
        set_border(self, ws, "A5:V5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 15
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 12
        ws.column_dimensions['S'].width = 12
        ws.column_dimensions['T'].width = 12
        ws.column_dimensions['U'].width = 12
        ws.column_dimensions['V'].width = 18

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Sector'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Provincia'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Distrito'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Ult. EESS His'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Documento'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Fecha Nacido'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Seguro'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['type'] == 'prematuro':
            ws.merge_cells('B2:V2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: SI-02.01: Niñas(os) prematuros y/o bajo peso al nacer, y niñas(os) de 4M de edad del departamento, sin diagnóstico de anemia; que reciben dosajes de hemoglobina y culminan la suplementación preventiva con hierro a los 6M de edad. - ' + nameMonth.upper() + ' ' + request.GET['anio']

            ws.merge_cells('B3:V3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: Dosaje Hemoglobina: 85018, 85018.01 - Suplementación: 99199.17, 99199.11 - Dx Anemmia: D500, D508, D509, D539, D649'

            ws.merge_cells('A5:V5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['I7'] = 'Visita 7D'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['J7'] = 'HB 30D'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['K7'] = 'Suple1'
            ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['L7'] = 'Suple2'
            ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['M7'] = 'Suple3'
            ws['M7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['M7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['M7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['M7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['N7'] = 'Dsje 3m'
            ws['N7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['N7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['N7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['N7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['O7'] = 'Suple4'
            ws['O7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['O7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['O7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['O7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['P7'] = 'Visita4'
            ws['P7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['P7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['P7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['P7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['Q7'] = 'Suple5'
            ws['Q7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['Q7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['Q7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['Q7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['R7'] = 'Visita5'
            ws['R7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['R7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['R7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['R7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['S7'] = 'Dsje 6m'
            ws['S7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['S7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['S7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['S7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['T7'] = 'TA'
            ws['T7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['T7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['T7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['T7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['U7'] = 'Cumple'
            ws['U7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['U7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['U7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['U7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            ws['V7'] = 'Programa'
            ws['V7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['V7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['V7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['V7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            if request.GET['sector'] == 'TODOS':
                if request.GET['prov'] == 'TODOS':
                    dataNom = si_0201_pr.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
                elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                    dataNom = si_0201_pr.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
                elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                    dataNom = si_0201_pr.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

            else:
                if request.GET['prov'] == 'TODOS':
                    dataNom = si_0201_pr.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
                elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                    dataNom = si_0201_pr.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
                elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                    dataNom = si_0201_pr.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for pqtrn in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                    ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                    ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                    ws.cell(row=cont, column=5).value = pqtrn['fields']['ult_eess']
                    ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=6).value = pqtrn['fields']['documento']
                    ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True, horizontal="center")
                    ws.cell(row=cont, column=7).value = pqtrn['fields']['fec_nac']
                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = pqtrn['fields']['seguro']
                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = pqtrn['fields']['visit7d']
                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = pqtrn['fields']['hb30d']
                    ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=11).value = pqtrn['fields']['suple1']
                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=12).value = pqtrn['fields']['suple2']
                    ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=13).value = pqtrn['fields']['suple3']
                    ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=14).value = pqtrn['fields']['dsje3m']
                    ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=15).value = pqtrn['fields']['suple4']
                    ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=16).value = pqtrn['fields']['visita4']
                    ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=17).value = pqtrn['fields']['suple5']
                    ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=18).value = pqtrn['fields']['visita5']
                    ws.cell(row=cont, column=18).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=19).value = pqtrn['fields']['dsje6m']
                    ws.cell(row=cont, column=19).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=20).value = pqtrn['fields']['ta6m']
                    ws.cell(row=cont, column=20).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=21).alignment = Alignment(horizontal="center")
                    if pqtrn['fields']['num'] == 1:
                        cumplen = '✔'
                        ws.cell(row=cont, column=21).font = Font(name='Calibri', size=10, color='00AC4E')
                    else:
                        cumplen = '✘'
                        ws.cell(row=cont, column=21).font = Font(name='Calibri', size=10, color='C00000')

                    ws.cell(row=cont, column=21).value = cumplen

                    ws.cell(row=cont, column=22).value = pqtrn['fields']['programa']
                    ws.cell(row=cont, column=22).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO NIÑOS PREMATUROS.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL NIÑOS'
            wb.save(response)
            return response

        else:
            ws.merge_cells('B2:P2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: SI-02.01: Niñas(os) prematuros y/o bajo peso al nacer, y niñas(os) de 4M de edad del departamento, sin diagnóstico de anemia; que reciben dosajes de hemoglobina y culminan la suplementación preventiva con hierro a los 6M de edad. - ' + nameMonth.upper() + ' ' + request.GET['anio']

            ws.merge_cells('B3:P3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: Dosaje Hemoglobina: 85018, 85018.01 - Suplementación: 99199.17, 99199.11 - Dx Anemmia: D500, D508, D509, D539, D649'

            ws.merge_cells('A5:P5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['I7'] = 'Suple4'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['J7'] = 'Visita4'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['K7'] = 'Suple5'
            ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['L7'] = 'Visita5'
            ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['M7'] = 'Dsje 6m'
            ws['M7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['M7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['M7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['M7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['N7'] = 'TA'
            ws['N7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['N7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['N7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['N7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

            ws['O7'] = 'Cumple'
            ws['O7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['O7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['O7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['O7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            ws['P7'] = 'Programa'
            ws['P7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['P7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['P7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['P7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            if request.GET['sector'] == 'TODOS':
                if request.GET['prov'] == 'TODOS':
                    dataNom = si_0201_sn.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
                elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                    dataNom = si_0201_sn.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
                elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                    dataNom = si_0201_sn.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

            else:
                if request.GET['prov'] == 'TODOS':
                    dataNom = si_0201_sn.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
                elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                    dataNom = si_0201_sn.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
                elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                    dataNom = si_0201_sn.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for pqtrn in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                    ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                    ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                    ws.cell(row=cont, column=5).value = pqtrn['fields']['ult_eess']
                    ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=6).value = pqtrn['fields']['documento']
                    ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True, horizontal="center")
                    ws.cell(row=cont, column=7).value = pqtrn['fields']['fec_nac']
                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = pqtrn['fields']['seguro']
                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = pqtrn['fields']['suple4']
                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = pqtrn['fields']['visita4']
                    ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=11).value = pqtrn['fields']['suple5']
                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=12).value = pqtrn['fields']['visita5']
                    ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=13).value = pqtrn['fields']['dsje6m']
                    ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=14).value = pqtrn['fields']['ta6m']
                    ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                    if pqtrn['fields']['num'] == 1:
                        cumplen = '✔'
                        ws.cell(row=cont, column=15).font = Font(name='Calibri', size=10, color='00AC4E')
                    else:
                        cumplen = '✘'
                        ws.cell(row=cont, column=15).font = Font(name='Calibri', size=10, color='C00000')

                    ws.cell(row=cont, column=15).value = cumplen

                    ws.cell(row=cont, column=16).value = pqtrn['fields']['programa']
                    ws.cell(row=cont, column=16).alignment = Alignment(wrap_text=True, horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO NIÑOS SUPLEMENTADOS DE 4 MESES.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL NIÑOS'
            wb.save(response)
            return response


class SI0202View(TemplateView):
    template_name = 'child/SI-02_02.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sector'] = Sector.objects.all()
        context['provincia'] = Provincia.objects.all()
        return context


class DistrictSI0202View(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListSI0202(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['sector'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                total = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
                cumplen = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
                cumplen = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
                cumplen = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['provincia'] == 'TODOS':
                total = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0202.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))


        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintSI0202(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:W2", "medium", "2F75B5")
        set_border(self, ws, "A3:W3", "medium", "366092")
        set_border(self, ws, "A5:W5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 15
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 12
        ws.column_dimensions['S'].width = 12
        ws.column_dimensions['T'].width = 12
        ws.column_dimensions['U'].width = 12
        ws.column_dimensions['V'].width = 8
        ws.column_dimensions['W'].width = 18

        ws.merge_cells('B2:W2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: SI-02.02: Niñas y niños de seis (06) meses de edad del departamento, con diagnóstico anemia, que reciben tres (03) dosajes de hemoglobina y culminan el tratamiento con hierro a los 12 meses de edad - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:W3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'CODIFICACION: Dosaje Hemoglobina: 85018, 85018.01 - Suplementación: 99199.17, 99199.11 - Dx Anemmia: D500, D508, D509, D539, D649'

        ws.merge_cells('A5:W5')
        ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Sector'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Provincia'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Distrito'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Establecimiento'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Ult. EESS His'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Documento'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Fecha Nacido'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['I7'] = 'Seguro'
        ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['J7'] = 'Dosaje 6M'
        ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['K7'] = 'Dx Anemia'
        ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['L7'] = 'Suple1'
        ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['M7'] = 'Suple2'
        ws['M7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M7'].fill = PatternFill(start_color='DCCBF5', end_color='DCCBF5', fill_type='solid')

        ws['N7'] = 'Dsj Ctrl1'
        ws['N7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N7'].fill = PatternFill(start_color='DCCBF5', end_color='DCCBF5', fill_type='solid')

        ws['O7'] = 'Suple3'
        ws['O7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O7'].fill = PatternFill(start_color='F5EFCB', end_color='F5EFCB', fill_type='solid')

        ws['P7'] = 'Dsj Ctrl2'
        ws['P7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P7'].fill = PatternFill(start_color='F5CBE7', end_color='F5CBE7', fill_type='solid')

        ws['Q7'] = 'Suple4'
        ws['Q7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q7'].fill = PatternFill(start_color='F5CBE7', end_color='F5CBE7', fill_type='solid')

        ws['R7'] = 'Suple5'
        ws['R7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['R7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R7'].fill = PatternFill(start_color='CBF5E4', end_color='CBF5E4', fill_type='solid')

        ws['S7'] = 'Suple6'
        ws['S7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['S7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S7'].fill = PatternFill(start_color='CBF5D5', end_color='CBF5D5', fill_type='solid')

        ws['T7'] = 'TA'
        ws['T7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['T7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T7'].fill = PatternFill(start_color='CBF5D5', end_color='CBF5D5', fill_type='solid')

        ws['U7'] = 'Dsje Ctrl'
        ws['U7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['U7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U7'].fill = PatternFill(start_color='CBF5D5', end_color='CBF5D5', fill_type='solid')

        ws['V7'] = 'Cumple'
        ws['V7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['V7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['V7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['V7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        ws['W7'] = 'Programa'
        ws['W7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['W7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['W7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['sector'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = si_0202.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = si_0202.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = si_0202.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = si_0202.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = si_0202.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = si_0202.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 8
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pqtrn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                ws.cell(row=cont, column=5).value = pqtrn['fields']['eess']
                ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=6).value = pqtrn['fields']['ult_eess']
                ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=7).value = pqtrn['fields']['documento']
                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).value = pqtrn['fields']['fec_nac']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = pqtrn['fields']['seguro']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = pqtrn['fields']['dsje6m']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).value = pqtrn['fields']['dx_anemia']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).value = pqtrn['fields']['suple1']
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).value = pqtrn['fields']['suple2']
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=14).value = pqtrn['fields']['dsje_ctrl1']
                ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=15).value = pqtrn['fields']['suple3']
                ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=16).value = pqtrn['fields']['dsje3']
                ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=17).value = pqtrn['fields']['suple4']
                ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=18).value = pqtrn['fields']['suple5']
                ws.cell(row=cont, column=18).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=19).value = pqtrn['fields']['suple6']
                ws.cell(row=cont, column=19).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=20).value = pqtrn['fields']['ta_1']
                ws.cell(row=cont, column=20).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=21).value = pqtrn['fields']['dsje_ctrl']
                ws.cell(row=cont, column=21).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=22).alignment = Alignment(horizontal="center")
                if pqtrn['fields']['num'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=22).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=22).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=22).value = cumplen
                ws.cell(row=cont, column=23).value = pqtrn['fields']['programa']
                ws.cell(row=cont, column=23).alignment = Alignment(wrap_text=True)

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO NIÑOS DE 6M CON DX ANEMIA.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL NIÑOS 6M CON DX ANEMIA'
        wb.save(response)
        return response


class SI0203View(TemplateView):
    template_name = 'child/SI-02_03.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sector'] = Sector.objects.all()
        context['provincia'] = Provincia.objects.all()
        return context


class DistrictSI0203View(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListSI0203(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['sector'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                total = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
                cumplen = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
                cumplen = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
                cumplen = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['provincia'] == 'TODOS':
                total = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0203.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))


        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintSI0203(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:U2", "medium", "2F75B5")
        set_border(self, ws, "A3:U3", "medium", "366092")
        set_border(self, ws, "A5:U5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 15
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 12
        ws.column_dimensions['S'].width = 12
        ws.column_dimensions['T'].width = 8
        ws.column_dimensions['U'].width = 20

        ws.merge_cells('B2:U2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: SI-02.03: Niñas y niños de seis (06) meses de edad del departamento, sin diagnóstico anemia, que reciben tres (03) dosajes de hemoglobina y culminan la suplementación preventiva con hierro a los 12 meses de edad - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:U3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'CODIFICACION: Dosaje Hemoglobina: 85018, 85018.01 - Suplementación: 99199.17, 99199.11'

        ws.merge_cells('A5:U5')
        ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Sector'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Provincia'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Distrito'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Establecimiento'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Ult. EESS His'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Documento'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Fecha Nacido'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['I7'] = 'Seguro'
        ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['J7'] = 'Dosaje 6M'
        ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['K7'] = 'Suple1'
        ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['L7'] = 'Suple2'
        ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L7'].fill = PatternFill(start_color='CBF5DF', end_color='CBF5DF', fill_type='solid')

        ws['M7'] = 'Suple3'
        ws['M7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M7'].fill = PatternFill(start_color='dce1f5', end_color='dce1f5', fill_type='solid')

        ws['N7'] = 'Dsje3'
        ws['N7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N7'].fill = PatternFill(start_color='dce1f5', end_color='dce1f5', fill_type='solid')

        ws['O7'] = 'Suple4'
        ws['O7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O7'].fill = PatternFill(start_color='F5CBE7', end_color='F5CBE7', fill_type='solid')

        ws['P7'] = 'Suple5'
        ws['P7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P7'].fill = PatternFill(start_color='CBF5E4', end_color='CBF5E4', fill_type='solid')

        ws['Q7'] = 'Suple6'
        ws['Q7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q7'].fill = PatternFill(start_color='E1CBF5', end_color='E1CBF5', fill_type='solid')

        ws['R7'] = 'TA'
        ws['R7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['R7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R7'].fill = PatternFill(start_color='CBF5D5', end_color='CBF5D5', fill_type='solid')

        ws['S7'] = 'Dsje 1A'
        ws['S7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['S7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S7'].fill = PatternFill(start_color='dce1f5', end_color='dce1f5', fill_type='solid')

        ws['T7'] = 'Cumple'
        ws['T7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['T7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['T7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        ws['U7'] = 'Programa'
        ws['U7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['U7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['sector'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = si_0203.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = si_0203.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = si_0203.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = si_0203.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = si_0203.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = si_0203.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 8
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pqtrn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                ws.cell(row=cont, column=5).value = pqtrn['fields']['eess']
                ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=6).value = pqtrn['fields']['ult_eess']
                ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=7).value = pqtrn['fields']['documento']
                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).value = pqtrn['fields']['fec_nac']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = pqtrn['fields']['seguro']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = pqtrn['fields']['dsje6m']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).value = pqtrn['fields']['suple1']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).value = pqtrn['fields']['suple2']
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).value = pqtrn['fields']['suple3']
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=14).value = pqtrn['fields']['dsje3_1']
                ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=15).value = pqtrn['fields']['suple4']
                ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=16).value = pqtrn['fields']['suple5']
                ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=17).value = pqtrn['fields']['suple6']
                ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=18).value = pqtrn['fields']['ta']
                ws.cell(row=cont, column=18).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=19).value = pqtrn['fields']['dsje12m']
                ws.cell(row=cont, column=19).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=20).alignment = Alignment(horizontal="center")
                if pqtrn['fields']['num'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=20).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=20).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=20).value = cumplen
                ws.cell(row=cont, column=21).value = pqtrn['fields']['programa']
                ws.cell(row=cont, column=21).alignment = Alignment(wrap_text=True)

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO NIÑOS DE 6M SIN DX ANEMIA.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL NIÑOS 6M SIN DX ANEMIA'
        wb.save(response)
        return response


class SI0401View(TemplateView):
    template_name = 'teen/SI-04_01.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sector'] = Sector.objects.all()
        context['provincia'] = Provincia.objects.all()
        return context


class DistrictSI0401View(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListSI0401(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['sector'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                total = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
                cumplen = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
                cumplen = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
                cumplen = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['provincia'] == 'TODOS':
                total = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = si_0401.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))


        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintSI0401(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:K2", "medium", "2F75B5")
        set_border(self, ws, "A3:K3", "medium", "366092")
        set_border(self, ws, "A5:K5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 32
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 8

        ws.merge_cells('B2:K2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: SI-04.01: Adolescentes mujeres de 12 a 17 años de edad, con dosaje de hemoglobina, en establecimientos de salud del primer y segundo nivel de atención (con población asignada) del Gobierno Regional. - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:K3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'CODIFICACION: Dosaje Hemoglobina: 85018, 85018.01'

        ws.merge_cells('A5:K5')
        ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Sector'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Provincia'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Distrito'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Establecimiento'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Categoria'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Documento'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Fecha Aten.'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['I7'] = 'Fecha HB'
        ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['J7'] = 'Suple'
        ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['K7'] = 'Cumple'
        ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['K7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        if request.GET['sector'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = si_0401.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = si_0401.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = si_0401.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = si_0401.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = si_0401.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = si_0401.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 8
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pqtrn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                ws.cell(row=cont, column=5).value = pqtrn['fields']['eess']
                ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=6).value = pqtrn['fields']['categoria']
                ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True, horizontal="center")
                ws.cell(row=cont, column=7).value = pqtrn['fields']['documento']
                ws.cell(row=cont, column=7).alignment = Alignment(wrap_text=True, horizontal="center")
                ws.cell(row=cont, column=8).value = pqtrn['fields']['atendida']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = pqtrn['fields']['fec_hb']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = pqtrn['fields']['suple']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                if pqtrn['fields']['num'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=11).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=11).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=11).value = cumplen

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO ADOLESCENTES CON DOSAJE HB.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL ADOLESCENTES CON DOSAJE'
        wb.save(response)
        return response


class VII0101View(TemplateView):
    template_name = 'pregnant/VII-01.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sector'] = Sector.objects.all()
        context['provincia'] = Provincia.objects.all()
        return context


class DistrictVII0101View(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListVII0101(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['sector'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                total = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
                cumplen = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
                cumplen = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
                cumplen = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['provincia'] == 'TODOS':
                total = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vii_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))


        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintVII0101(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:T2", "medium", "2F75B5")
        set_border(self, ws, "A3:T3", "medium", "366092")
        set_border(self, ws, "A5:T5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 32
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 12
        ws.column_dimensions['S'].width = 8
        ws.column_dimensions['T'].width = 18

        ws.merge_cells('B2:T2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: VII-01: Gestantes atendidas en eess de salud del I y II nivel de atención (con población asignada) del Gobierno Regional, con diagnóstico de violencia, que reciben un paquete mínimo de intervenciones terapeuticas especializadas. - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:T3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'CODIFICACION: Dx Violencia: 96150, 96150.01 - Tmz Positivo: R456 - T74.0: Negligencia y abandono - T74.1: Abuso físico - T74.2 Abuso sexual - T74.3 Abuso psicológico - T74.8 Otros síndromes del maltrato (formas mixtas) - T74.9 Síndrome de maltrato, no especificado - Y04 hasta Y08: Agresiones'

        ws.merge_cells('A5:T5')
        ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Sector'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Provincia'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Distrito'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Eess Apn'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Eess Dx'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Documento'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Diagnóstico'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['I7'] = 'Dx 3m'
        ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I7'].fill = PatternFill(start_color='DCCBF5', end_color='DCCBF5', fill_type='solid')

        ws['J7'] = 'Dx 6m'
        ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J7'].fill = PatternFill(start_color='F5CBD7', end_color='F5CBD7', fill_type='solid')

        ws['K7'] = 'Csm1'
        ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K7'].fill = PatternFill(start_color='F5D4CB', end_color='F5D4CB', fill_type='solid')

        ws['L7'] = 'Csm2'
        ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L7'].fill = PatternFill(start_color='F5D4CB', end_color='F5D4CB', fill_type='solid')

        ws['M7'] = 'Psicol. 1'
        ws['M7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M7'].fill = PatternFill(start_color='F5EFCB', end_color='F5EFCB', fill_type='solid')

        ws['N7'] = 'Psicol. 2'
        ws['N7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N7'].fill = PatternFill(start_color='F5EFCB', end_color='F5EFCB', fill_type='solid')

        ws['O7'] = 'Psicol. 3'
        ws['O7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O7'].fill = PatternFill(start_color='F5EFCB', end_color='F5EFCB', fill_type='solid')

        ws['P7'] = 'Psicol. 4'
        ws['P7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P7'].fill = PatternFill(start_color='F5EFCB', end_color='F5EFCB', fill_type='solid')

        ws['Q7'] = 'Psicol. 5'
        ws['Q7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q7'].fill = PatternFill(start_color='F5EFCB', end_color='F5EFCB', fill_type='solid')

        ws['R7'] = 'Psicol. 6'
        ws['R7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['R7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R7'].fill = PatternFill(start_color='F5EFCB', end_color='F5EFCB', fill_type='solid')

        ws['S7'] = 'Cumple'
        ws['S7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['S7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['S7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        ws['T7'] = 'Programa'
        ws['T7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['T7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['sector'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = vii_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = vii_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = vii_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = vii_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = vii_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = vii_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 8
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pqtrn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                ws.cell(row=cont, column=5).value = pqtrn['fields']['eess_apn']
                ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=6).value = pqtrn['fields']['eess_dx']
                ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=7).value = pqtrn['fields']['documento']
                ws.cell(row=cont, column=7).alignment = Alignment(wrap_text=True, horizontal="center")
                ws.cell(row=cont, column=8).value = pqtrn['fields']['diagnostico']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = pqtrn['fields']['dx_3m']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = pqtrn['fields']['dx_6m']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).value = pqtrn['fields']['csm1']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).value = pqtrn['fields']['csm2']
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).value = pqtrn['fields']['psicologia1']
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=14).value = pqtrn['fields']['psicologia2']
                ws.cell(row=cont, column=14).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=15).value = pqtrn['fields']['psicologia3']
                ws.cell(row=cont, column=15).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=16).value = pqtrn['fields']['psicologia4']
                ws.cell(row=cont, column=16).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=17).value = pqtrn['fields']['psicologia5']
                ws.cell(row=cont, column=17).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=18).value = pqtrn['fields']['psicologia6']
                ws.cell(row=cont, column=18).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=19).alignment = Alignment(horizontal="center")
                if pqtrn['fields']['num'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=19).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=19).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=19).value = cumplen
                ws.cell(row=cont, column=20).value = pqtrn['fields']['programa']
                ws.cell(row=cont, column=20).alignment = Alignment(wrap_text=True, horizontal="center")

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO GESTANTES CON DX VIOLENCIA QUE RECIBEN PQT.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL GEST. DX VIOLENCIA QUE RECIBEN PQT'
        wb.save(response)
        return response


class VI0101View(TemplateView):
    template_name = 'pregnant/VI-01_01.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sector'] = Sector.objects.all()
        context['provincia'] = Provincia.objects.all()
        return context


class DistrictVI0101View(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListVI0101(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['sector'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                total = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['provincia'] == 'TODOS':
                total = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0101.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))


        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintVI0101(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:L2", "medium", "2F75B5")
        set_border(self, ws, "A3:L3", "medium", "366092")
        set_border(self, ws, "A5:L5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 32
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 18

        ws.merge_cells('B2:L2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: VI-01: Gestantes atendidas en establecimientos de salud del primer y segundo nivel de atención (con población asignada) del Gobierno Regional, que durante el embarazo le aplicaron la ficha de detección de violencia contra la mujer. - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:L3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'CODIFICACION: Dx Violencia: 96150, 96150.01'

        ws.merge_cells('A5:L5')
        ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Sector'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Provincia'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Distrito'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Eess'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Categoria'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Documento'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Trimestre'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['I7'] = 'Fecha Apn'
        ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['J7'] = 'Tmz'
        ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J7'].fill = PatternFill(start_color='DCCBF5', end_color='DCCBF5', fill_type='solid')

        ws['K7'] = 'Cumple'
        ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['K7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        ws['L7'] = 'Programa'
        ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['sector'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = vi_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = vi_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = vi_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = vi_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = vi_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = vi_0101.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 8
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pqtrn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                ws.cell(row=cont, column=5).value = pqtrn['fields']['eess']
                ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=6).value = pqtrn['fields']['categoria']
                ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True, horizontal="center")
                ws.cell(row=cont, column=7).value = pqtrn['fields']['documento']
                ws.cell(row=cont, column=7).alignment = Alignment(wrap_text=True, horizontal="center")
                ws.cell(row=cont, column=8).value = pqtrn['fields']['trimestre']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = pqtrn['fields']['fecha_apn']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = pqtrn['fields']['tmz']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                if pqtrn['fields']['num'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=11).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=11).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=11).value = cumplen
                ws.cell(row=cont, column=12).value = pqtrn['fields']['programa']
                ws.cell(row=cont, column=12).alignment = Alignment(wrap_text=True, horizontal="center")

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO GEST. QUE LE APLICARON FICHA DX VIOLENCIA.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL GEST. APLICADOS FICHA DX VIOLENCIA'
        wb.save(response)
        return response


class VI0102View(TemplateView):
    template_name = 'pregnant/VI-01_02.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sector'] = Sector.objects.all()
        context['provincia'] = Provincia.objects.all()
        return context


class DistrictVI0102View(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class ListVI0102(View):
    def post(self, request, *args, **kwargs):
        json_data4 = []
        if request.POST['sector'] == 'TODOS':
            if request.POST['provincia'] == 'TODOS':
                total = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        else:
            if request.POST['provincia'] == 'TODOS':
                total = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] == 'TODOS':
                total = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_prov=request.POST['provincia'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.POST['provincia'] != 'TODOS' and request.POST['distrito'] != 'TODOS':
                total = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(total=Sum('den'))['total']
                cumplen = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).aggregate(cumplen=Sum('num'))['cumplen']
                dataTotal = { 'total': total if total is not None else 0, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if (total != 0) and (total is not None) else 0 }
                dataProv = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('provincia').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataDist = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).values('distrito').annotate(denominador=Sum('den'),
                            numerador=Sum('num'), avance=(ExpressionWrapper( Cast(Sum('num'), FloatField()) / Cast(Sum('den'), FloatField()) * 100,
                            output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                dataNom = vi_0102.objects.filter(anio=request.POST['anio'], mes=request.POST['mes'], cod_dist=request.POST['distrito'], cod_sector=request.POST['sector']).order_by('provincia', 'distrito')
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))


        json_data4.append(dataTotal)
        json_data4.append(list(dataProv))
        json_data4.append(list(dataDist))
        json_data4.append(dataNom)

        return HttpResponse(json.dumps(json_data4), content_type='application/json')


class PrintVI0102(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        locale.setlocale(locale.LC_ALL, "C")
        nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:N2", "medium", "2F75B5")
        set_border(self, ws, "A3:N3", "medium", "366092")
        set_border(self, ws, "A5:N5", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        ws.row_dimensions[2].height = 32
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 8
        ws.column_dimensions['N'].width = 18

        ws.merge_cells('B2:N2')
        ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO - DEIT: VI-02: Gestantes atendidas en establecimientos de salud del primer y segundo nivel de atención (con población asignada) del Gobierno Regional, que cuentan tamizaje positivo de violencia contra la mujer. - ' + nameMonth.upper() + ' ' + request.GET['anio']

        ws.merge_cells('B3:N3')
        ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['B3'] = 'CODIFICACION: Dx Violencia: 96150, 96150.01 - Tmz Positivo: R456'

        ws.merge_cells('A5:N5')
        ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A7'] = '#'
        ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['B7'] = 'Sector'
        ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['C7'] = 'Provincia'
        ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['D7'] = 'Distrito'
        ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['E7'] = 'Eess'
        ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['F7'] = 'Categoria'
        ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['G7'] = 'Documento'
        ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['H7'] = 'Fecha Nacido'
        ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        ws['I7'] = 'Gest Tmz'
        ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['J7'] = 'Trimestre'
        ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['K7'] = 'Tmz'
        ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['L7'] = 'Sospecha'
        ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L7'].fill = PatternFill(start_color='CBEFF5', end_color='CBEFF5', fill_type='solid')

        ws['M7'] = 'Cumple'
        ws['M7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['M7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        ws['N7'] = 'Programa'
        ws['N7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N7'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['N7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

        if request.GET['sector'] == 'TODOS':
            if request.GET['prov'] == 'TODOS':
                dataNom = vi_0102.objects.filter(anio=request.GET['anio'], mes=request.GET['mes']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = vi_0102.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = vi_0102.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_dist=request.GET['dist']).order_by('distrito')

        else:
            if request.GET['prov'] == 'TODOS':
                dataNom = vi_0102.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = vi_0102.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS':
                dataNom = vi_0102.objects.filter(anio=request.GET['anio'], mes=request.GET['mes'], cod_sector=request.GET['sector'], cod_dist=request.GET['dist']).order_by('distrito')

        dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 8
        cant = len(dataNom)
        num=1
        if cant > 0:
            for pqtrn in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=2).value = pqtrn['fields']['sector']
                ws.cell(row=cont, column=3).value = pqtrn['fields']['provincia']
                ws.cell(row=cont, column=4).value = pqtrn['fields']['distrito']
                ws.cell(row=cont, column=5).value = pqtrn['fields']['eess']
                ws.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                ws.cell(row=cont, column=6).value = pqtrn['fields']['categoria']
                ws.cell(row=cont, column=6).alignment = Alignment(wrap_text=True, horizontal="center")
                ws.cell(row=cont, column=7).value = pqtrn['fields']['documento']
                ws.cell(row=cont, column=7).alignment = Alignment(wrap_text=True, horizontal="center")
                ws.cell(row=cont, column=8).value = pqtrn['fields']['fec_nac']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = pqtrn['fields']['gest_tmz']
                ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=10).value = pqtrn['fields']['trimestre']
                ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=11).value = pqtrn['fields']['tmz']
                ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=12).value = pqtrn['fields']['sospecha']
                ws.cell(row=cont, column=12).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                if pqtrn['fields']['num'] == 1:
                    cumplen = '✔'
                    ws.cell(row=cont, column=13).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=13).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=13).value = cumplen
                ws.cell(row=cont, column=14).value = pqtrn['fields']['programa']
                ws.cell(row=cont, column=14).alignment = Alignment(wrap_text=True, horizontal="center")

                cont = cont+1
                num = num+1

        nombre_archivo = "DEIT_PASCO GEST. QUE CUENTAN CON TMZ POSITIVO.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL GEST. CON DX VIOLENCIA'
        wb.save(response)
        return response

