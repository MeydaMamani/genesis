from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, get_object_or_404
from django.views.generic import TemplateView, View
from django.core import serializers
from django.http import JsonResponse, HttpResponse, QueryDict
from django.db.models import Case, When, IntegerField, FloatField, ExpressionWrapper, Q, F, Sum, Count, IntegerField, Avg, Value, DecimalField
from django.db.models.functions import Cast, Round
import json
from datetime import date, datetime

from .models import dit001_ac_n, mat002_vg_n, mat002_vg_cr, mat002_vg_c, mat002_dc_n, dit001_fng_n, mat002_vp_n, mat002_vp_cr, tbcvih016_tbc_n, tbcvih016_tbc_cr, tbcvih016_tbc_c, tbcvih016_vih_n, tbcvih016_vih_cr, tbcvih016_vih_c, met017_met_n, met017_met_cr, cc
from apps.main.models import Sector, Provincia, Distrito, Establecimiento

# library excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color

import locale
import datetime

class PromsaView(TemplateView):
    template_name = 'index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['provincia'] = Provincia.objects.exclude(codigo__in=['00'])
        return context


class Districts(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Distrito.objects.filter(prov_id=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(dist, content_type='application/json')


class EESS(View):
    def get(self, request, *args, **kwargs):
        eess = serializers.serialize('json', Establecimiento.objects.filter(dist_id=request.GET['id'], sector_id=7), indent=2, use_natural_foreign_keys=True)
        return HttpResponse(eess, content_type='application/json')


class PrintPromsa(View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active

        # locale.setlocale(locale.LC_ALL, "C")
        # nameMonth = date(1900, int(request.GET['mes']), 1).strftime('%B')

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        img = Image('static/img/logoPrint.png')
        ws.merge_cells('A2:A3')
        ws.add_image(img, 'A2')

        if request.GET['tipo'] == 'dit001_acn':
            set_border(self, ws, "A2:I2", "medium", "2F75B5")
            set_border(self, ws, "A3:I3", "medium", "366092")
            set_border(self, ws, "A5:I5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 32
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 35
            ws.column_dimensions['H'].width = 32

            ws.merge_cells('B2:I2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: 1001 - DIT- Actores sociales capacitados para la promoción del cuidado infantil, lactancia materna exclusiva - ' + request.GET['anio']

            ws.merge_cells('B3:I3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:I5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Mes'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Participantes'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Taller'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Sub Producto'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Reg'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            if request.GET['prov'] == 'TODOS':
                dataNom = dit001_ac_n.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = dit001_ac_n.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = dit001_ac_n.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = dit001_ac_n.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for dit01 in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = dit01['fields']['provincia']
                    ws.cell(row=cont, column=3).value = dit01['fields']['distrito']
                    ws.cell(row=cont, column=4).value = dit01['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = dit01['fields']['mes']
                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = dit01['fields']['participante']
                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = dit01['fields']['taller']
                    ws.cell(row=cont, column=7).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=8).value = dit01['fields']['subproduct']
                    ws.cell(row=cont, column=8).alignment = Alignment( wrap_text=True)
                    ws.cell(row=cont, column=9).value = dit01['fields']['reg']
                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO ACTORES SOCIALES CAPACITADOS.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            wb.save(response)
            return response

        elif request.GET['tipo'] == 'dit001_fng':
            set_border(self, ws, "A2:k2", "medium", "2F75B5")
            set_border(self, ws, "A3:k3", "medium", "366092")
            set_border(self, ws, "A5:k5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 24
            ws.column_dimensions['E'].width = 5
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 5
            ws.column_dimensions['I'].width = 5
            ws.column_dimensions['J'].width = 36
            ws.column_dimensions['K'].width = 5

            ws.merge_cells('B2:k2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: 1001 - DIT - Familias con niños menores de 12 meses y gestantes reciben acompañamiento a través de sesiones demostrativas  - ' + request.GET['anio']

            ws.merge_cells('B3:k3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:k5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Mes'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Documento'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Fecha Aten.'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Edad Reg'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Tipo Edad'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['I7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['J7'] = 'Sub Producto'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['K7'] = 'Reg'
            ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            if request.GET['prov'] == 'TODOS':
                dataNom = dit001_fng_n.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = dit001_fng_n.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = dit001_fng_n.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = dit001_fng_n.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for dit01 in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = dit01['fields']['provincia']
                    ws.cell(row=cont, column=3).value = dit01['fields']['distrito']
                    ws.cell(row=cont, column=4).value = dit01['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = dit01['fields']['mes']
                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = dit01['fields']['documento']
                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = dit01['fields']['fec_atencion']
                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = dit01['fields']['edad_reg']
                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = dit01['fields']['tedad']
                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = dit01['fields']['subproduct']
                    ws.cell(row=cont, column=10).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=11).value = dit01['fields']['reg']
                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO FAMILIAS CON MENORES DE 12M Y GEST. QUE RECIBEN SESIONES DEMOST..xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            wb.save(response)
            return response

        elif request.GET['tipo'] == 'mat002vg':
            sheet2 = wb.create_sheet('CRUCE')
            sheet3 = wb.create_sheet('CONTEO')

            if request.GET['prov'] == 'TODOS':
                dataNom = mat002_vg_n.objects.filter(anio=request.GET['anio']).order_by('provincia')
                dataCruce = mat002_vg_cr.objects.filter(anio=request.GET['anio']).order_by('provincia')
                dataCont = mat002_vg_c.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = mat002_vg_n.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
                dataCruce = mat002_vg_cr.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
                dataCont = mat002_vg_c.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = mat002_vg_n.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCruce = mat002_vg_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCont = mat002_vg_c.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = mat002_vg_n.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
                dataCruce = mat002_vg_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCont = mat002_vg_c.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
            dataCruce = json.loads(serializers.serialize('json', dataCruce, indent=2, use_natural_foreign_keys=True))
            dataCont = json.loads(serializers.serialize('json', dataCont, indent=2, use_natural_foreign_keys=True))

            set_border(self, ws, "A2:L2", "medium", "2F75B5")
            set_border(self, ws, "A3:L3", "medium", "366092")
            set_border(self, ws, "A5:L5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 24
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['K'].width = 9
            ws.column_dimensions['L'].width = 32

            ws.merge_cells('B2:L2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: 002 - MATERNO NEONATAL - Familias de la gestante que recibe consejería en el hogar - ' + request.GET['anio']

            ws.merge_cells('B3:L3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:L5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Documento'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Mes'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Fecha Atención'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Lab'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Reg'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['J7'] = 'Trimestre'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['K7'] = 'Trazador'
            ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['L7'] = 'Sub Producto'
            ws['L7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['L7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['L7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for mat002 in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = mat002['fields']['provincia']
                    ws.cell(row=cont, column=3).value = mat002['fields']['distrito']
                    ws.cell(row=cont, column=4).value = mat002['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = mat002['fields']['documento']
                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = mat002['fields']['mes']
                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = mat002['fields']['fec_atencion']
                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = mat002['fields']['lab']
                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = mat002['fields']['reg']
                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = mat002['fields']['trimestre']
                    ws.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=11).value = mat002['fields']['trazador']
                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=12).value = mat002['fields']['subproduct']
                    ws.cell(row=cont, column=12).alignment = Alignment(wrap_text=True)

                    cont = cont+1
                    num = num+1


            sheet2.row_dimensions[2].height = 23
            sheet2.column_dimensions['A'].width = 7
            sheet2.column_dimensions['B'].width = 24
            sheet2.column_dimensions['C'].width = 24
            sheet2.column_dimensions['D'].width = 32
            sheet2.column_dimensions['E'].width = 12
            sheet2.column_dimensions['F'].width = 12
            sheet2.column_dimensions['G'].width = 12
            sheet2.column_dimensions['H'].width = 12
            sheet2.column_dimensions['I'].width = 7
            sheet2.column_dimensions['J'].width = 12
            sheet2.column_dimensions['K'].width = 12

            set_border(self, sheet2, "A2:K2", "medium", "2F75B5")

            sheet2.merge_cells('A2:K2')
            sheet2['A2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            sheet2['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sheet2['A2'] = 'DIRESA PASCO - DEIT: 002 - MATERNO NEONATAL - Familias de la gestante que recibe consejería en el hogar (CRUCE) - ' + request.GET['anio']

            sheet2['A4'] = '#'
            sheet2['A4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['A4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['A4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['A4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['B4'] = 'Provincia'
            sheet2['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['B4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['B4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['C4'] = 'Distrito'
            sheet2['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['C4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['C4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['C4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['D4'] = 'Establecimiento'
            sheet2['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['D4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['D4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['D4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['E4'] = 'Documento'
            sheet2['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['E4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['E4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['E4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['F4'] = 'Primer Trim'
            sheet2['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['F4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['F4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['F4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['G4'] = 'Segundo Trim'
            sheet2['G4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['G4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['G4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['G4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['H4'] = 'Tercer Trim'
            sheet2['H4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['H4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['H4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['H4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['I4'] = 'Trazador'
            sheet2['I4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['I4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['I4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['I4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['J4'] = 'FUR'
            sheet2['J4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['J4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['J4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['J4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['K4'] = 'FPP'
            sheet2['K4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['K4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['K4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['K4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 5
            cant = len(dataCruce)
            num=1
            if cant > 0:
                for mat002_cr in dataCruce:
                    sheet2.cell(row=cont, column=1).value = num
                    sheet2.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=2).value = mat002_cr['fields']['provincia']
                    sheet2.cell(row=cont, column=3).value = mat002_cr['fields']['distrito']
                    sheet2.cell(row=cont, column=4).value = mat002_cr['fields']['eess']
                    sheet2.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    sheet2.cell(row=cont, column=5).value = mat002_cr['fields']['documento']
                    sheet2.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=6).value = mat002_cr['fields']['primer_trim']
                    sheet2.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=7).value = mat002_cr['fields']['segundo_trim']
                    sheet2.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=8).value = mat002_cr['fields']['tercer_trim']
                    sheet2.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=9).value = mat002_cr['fields']['trazador']
                    sheet2.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=10).value = mat002_cr['fields']['fur']
                    sheet2.cell(row=cont, column=10).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=11).value = mat002_cr['fields']['fpp']
                    sheet2.cell(row=cont, column=11).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1


            sheet3.row_dimensions[2].height = 36
            sheet3.column_dimensions['A'].width = 7
            sheet3.column_dimensions['B'].width = 24
            sheet3.column_dimensions['C'].width = 40
            sheet3.column_dimensions['D'].width = 10
            sheet3.column_dimensions['E'].width = 10
            sheet3.column_dimensions['F'].width = 15

            set_border(self, sheet3, "A2:F2", "medium", "2F75B5")

            sheet3.merge_cells('A2:F2')
            sheet3['A2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            sheet3['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sheet3['A2'] = 'DIRESA PASCO - DEIT: 002 - MATERNO NEONATAL - Familias de la gestante que recibe consejería en el hogar (CONSOLIDADO) - ' + request.GET['anio']

            sheet3['A4'] = '#'
            sheet3['A4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['A4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['A4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['A4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['B4'] = 'Provincia'
            sheet3['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['B4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet3['B4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['C4'] = 'Distrito'
            sheet3['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['C4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['C4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['C4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['D4'] = 'Total'
            sheet3['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['D4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['D4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['D4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['E4'] = 'Cumplen'
            sheet3['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['E4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['E4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet3['E4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['F4'] = 'Avance'
            sheet3['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['F4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['F4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['F4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            cont = 5
            cant = len(dataCont)
            num=1
            if cant > 0:
                for mat002_c in dataCont:
                    sheet3.cell(row=cont, column=1).value = num
                    sheet3.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    sheet3.cell(row=cont, column=2).value = mat002_c['fields']['provincia']
                    sheet3.cell(row=cont, column=3).value = mat002_c['fields']['distrito']
                    sheet3.cell(row=cont, column=4).value = mat002_c['fields']['den']
                    sheet3.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
                    sheet3.cell(row=cont, column=5).value = mat002_c['fields']['num']
                    sheet3.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    avance = round((mat002_c['fields']['num']/mat002_c['fields']['den'])*100, 1) if (mat002_c['fields']['den'] != 0) and (mat002_c['fields']['den'] is not None) else 0
                    sheet3.cell(row=cont, column=6).value = str(avance) + ' %'
                    sheet3.cell(row=cont, column=6).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO MATERNO NEONATAL - FAMILIAS GESTANTE QUE RECIB. CONSEJ..xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            sheet2.title = 'CRUCE'
            sheet3.title = 'CONSOLIDADO'
            wb.save(response)
            return response

        elif request.GET['tipo'] == 'mat002dc':
            set_border(self, ws, "A2:I2", "medium", "2F75B5")
            set_border(self, ws, "A3:I3", "medium", "366092")
            set_border(self, ws, "A5:I5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 5
            ws.column_dimensions['F'].width = 13
            ws.column_dimensions['G'].width = 45
            ws.column_dimensions['H'].width = 68

            ws.merge_cells('B2:I2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: 002- MATERNO NEONATAL - Docentes capacitados realizan educación sexual integral desde la intuición educativa  - ' + request.GET['anio']

            ws.merge_cells('B3:I3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:I5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Mes'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Eje Temático'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Taller'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Sub Producto'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Reg'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            if request.GET['prov'] == 'TODOS':
                dataNom = mat002_dc_n.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = mat002_dc_n.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = mat002_dc_n.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = mat002_dc_n.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for dit01 in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = dit01['fields']['provincia']
                    ws.cell(row=cont, column=3).value = dit01['fields']['distrito']
                    ws.cell(row=cont, column=4).value = dit01['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = dit01['fields']['mes']
                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = dit01['fields']['eje']
                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = dit01['fields']['taller']
                    ws.cell(row=cont, column=7).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=8).value = dit01['fields']['subproduct']
                    ws.cell(row=cont, column=8).alignment = Alignment( wrap_text=True)
                    ws.cell(row=cont, column=9).value = dit01['fields']['reg']
                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO MATERNO NEONATAL - DOCENTES CAPACITADOS.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            wb.save(response)
            return response

        elif request.GET['tipo'] == 'mat002vp':
            sheet2 = wb.create_sheet('CRUCE')

            if request.GET['prov'] == 'TODOS':
                dataNom = mat002_vp_n.objects.filter(anio=request.GET['anio']).order_by('provincia')
                dataCruce = mat002_vp_cr.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = mat002_vp_n.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
                dataCruce = mat002_vp_cr.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = mat002_vp_n.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCruce = mat002_vp_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = mat002_vp_n.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
                dataCruce = mat002_vp_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
            dataCruce = json.loads(serializers.serialize('json', dataCruce, indent=2, use_natural_foreign_keys=True))

            set_border(self, ws, "A2:K2", "medium", "2F75B5")
            set_border(self, ws, "A3:K3", "medium", "366092")
            set_border(self, ws, "A5:K5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 24
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['J'].width = 25
            ws.column_dimensions['K'].width = 7

            ws.merge_cells('B2:K2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: 002- MATERNO NEONATAL - Familias de la puerpera que recibe consejería en el hogar - ' + request.GET['anio']

            ws.merge_cells('B3:K3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:K5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Documento'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Mes'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Fecha Atención'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Visita'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Trazador'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['J7'] = 'Sub Producto'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['K7'] = 'Reg'
            ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for mat002vpn in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = mat002vpn['fields']['provincia']
                    ws.cell(row=cont, column=3).value = mat002vpn['fields']['distrito']
                    ws.cell(row=cont, column=4).value = mat002vpn['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = mat002vpn['fields']['documento']
                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = mat002vpn['fields']['mes']
                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = mat002vpn['fields']['fec_atencion']
                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = mat002vpn['fields']['visita']
                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = mat002vpn['fields']['trazador']
                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = mat002vpn['fields']['subproduct']
                    ws.cell(row=cont, column=10).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=11).value = mat002vpn['fields']['reg']
                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1


            sheet2.row_dimensions[2].height = 23
            sheet2.column_dimensions['A'].width = 7
            sheet2.column_dimensions['B'].width = 24
            sheet2.column_dimensions['C'].width = 24
            sheet2.column_dimensions['D'].width = 32
            sheet2.column_dimensions['E'].width = 12
            sheet2.column_dimensions['F'].width = 12
            sheet2.column_dimensions['G'].width = 12
            sheet2.column_dimensions['H'].width = 12

            set_border(self, sheet2, "A2:H2", "medium", "2F75B5")

            sheet2.merge_cells('A2:H2')
            sheet2['A2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            sheet2['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sheet2['A2'] = 'DIRESA PASCO - DEIT: 002- MATERNO NEONATAL - Familias de la puerpera que recibe consejería en el hogar (CRUCE) - ' + request.GET['anio']

            sheet2['A4'] = '#'
            sheet2['A4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['A4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['A4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['A4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['B4'] = 'Provincia'
            sheet2['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['B4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['B4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['C4'] = 'Distrito'
            sheet2['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['C4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['C4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['C4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['D4'] = 'Establecimiento'
            sheet2['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['D4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['D4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['D4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['E4'] = 'Documento'
            sheet2['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['E4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['E4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['E4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['F4'] = 'Visita1'
            sheet2['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['F4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['F4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['F4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['G4'] = 'Visita2'
            sheet2['G4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['G4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['G4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['G4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['H4'] = 'Visita3'
            sheet2['H4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['H4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['H4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['H4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 5
            cant = len(dataCruce)
            num=1
            if cant > 0:
                for mat002vpncr in dataCruce:
                    sheet2.cell(row=cont, column=1).value = num
                    sheet2.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=2).value = mat002vpncr['fields']['provincia']
                    sheet2.cell(row=cont, column=3).value = mat002vpncr['fields']['distrito']
                    sheet2.cell(row=cont, column=4).value = mat002vpncr['fields']['eess']
                    sheet2.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    sheet2.cell(row=cont, column=5).value = mat002vpncr['fields']['documento']
                    sheet2.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=6).value = mat002vpncr['fields']['visita1']
                    sheet2.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=7).value = mat002vpncr['fields']['visita2']
                    sheet2.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=8).value = mat002vpncr['fields']['visita3']
                    sheet2.cell(row=cont, column=8).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO MATERNO NEONATAL - FAMILIAS DE LA PUERPERA QUE RECIB. CONSEJ.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            sheet2.title = 'CRUCE'
            wb.save(response)
            return response

        elif request.GET['tipo'] == 'tbcvih_tbc':
            sheet2 = wb.create_sheet('CRUCE')
            sheet3 = wb.create_sheet('CONTEO')

            if request.GET['prov'] == 'TODOS':
                dataNom = tbcvih016_tbc_n.objects.filter(anio=request.GET['anio']).order_by('provincia')
                dataCruce = tbcvih016_tbc_cr.objects.filter(anio=request.GET['anio']).order_by('provincia')
                dataCont = tbcvih016_tbc_c.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = tbcvih016_tbc_n.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
                dataCruce = tbcvih016_tbc_cr.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
                dataCont = tbcvih016_tbc_c.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = tbcvih016_tbc_n.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCruce = tbcvih016_tbc_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCont = tbcvih016_tbc_c.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = tbcvih016_tbc_n.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
                dataCruce = tbcvih016_tbc_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCont = tbcvih016_tbc_c.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
            dataCruce = json.loads(serializers.serialize('json', dataCruce, indent=2, use_natural_foreign_keys=True))
            dataCont = json.loads(serializers.serialize('json', dataCont, indent=2, use_natural_foreign_keys=True))

            set_border(self, ws, "A2:I2", "medium", "2F75B5")
            set_border(self, ws, "A3:I3", "medium", "366092")
            set_border(self, ws, "A5:I5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 36
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 50
            ws.column_dimensions['J'].width = 14

            ws.merge_cells('B2:J2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: 0016 TBC-VIH/SIDA - TBC - Familias que reciben consejería para promover prácticas y entornos saludables para contribuir a la disminución de la Tuberculosis y VIH/SIDA - ' + request.GET['anio']

            ws.merge_cells('B3:J3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:J5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Documento'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Mes'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Financiador'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Visita'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Sub Producto'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['J7'] = 'Reg'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for tbc in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = tbc['fields']['provincia']
                    ws.cell(row=cont, column=3).value = tbc['fields']['distrito']
                    ws.cell(row=cont, column=4).value = tbc['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = tbc['fields']['documento']
                    ws.cell(row=cont, column=6).value = tbc['fields']['mes']
                    ws.cell(row=cont, column=7).value = tbc['fields']['financiador']
                    ws.cell(row=cont, column=8).value = tbc['fields']['visita']
                    ws.cell(row=cont, column=9).value = tbc['fields']['subproduct']
                    ws.cell(row=cont, column=9).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=10).value = tbc['fields']['reg']

                    cont = cont+1
                    num = num+1

            sheet2.row_dimensions[2].height = 36
            sheet2.column_dimensions['A'].width = 7
            sheet2.column_dimensions['B'].width = 24
            sheet2.column_dimensions['C'].width = 24
            sheet2.column_dimensions['D'].width = 10
            sheet2.column_dimensions['E'].width = 15
            sheet2.column_dimensions['F'].width = 26
            sheet2.column_dimensions['G'].width = 12
            sheet2.column_dimensions['H'].width = 26
            sheet2.column_dimensions['I'].width = 12
            sheet2.column_dimensions['J'].width = 8

            set_border(self, sheet2, "A2:J2", "medium", "2F75B5")

            sheet2.merge_cells('A2:J2')
            sheet2['A2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            sheet2['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sheet2['A2'] = 'DIRESA PASCO - DEIT: 0016 TBC-VIH/SIDA - TBC-Familias que reciben consejería para promover prácticas y entornos saludables para contribuir a la disminución de la Tuberculosis y VIH/SIDA (CRUCE) - ' + request.GET['anio']

            sheet2['A4'] = '#'
            sheet2['A4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['A4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['A4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['A4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['B4'] = 'Provincia'
            sheet2['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['B4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['B4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['C4'] = 'Distrito'
            sheet2['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['C4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['C4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['C4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['D4'] = 'Documento'
            sheet2['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['D4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['D4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['D4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['E4'] = 'Financiador'
            sheet2['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['E4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['E4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['E4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['F4'] = 'EESS Vista1'
            sheet2['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['F4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['F4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['F4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['G4'] = 'Visita1'
            sheet2['G4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['G4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['G4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['G4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['H4'] = 'EESS Vista2'
            sheet2['H4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['H4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['H4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['H4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['I4'] = 'Visita2'
            sheet2['I4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['I4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['I4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['I4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['J4'] = 'Días Visit.'
            sheet2['J4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['J4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['J4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['J4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 5
            cant = len(dataCruce)
            num=1
            if cant > 0:
                for tbc_cr in dataCruce:
                    sheet2.cell(row=cont, column=1).value = num
                    sheet2.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=2).value = tbc_cr['fields']['provincia']
                    sheet2.cell(row=cont, column=3).value = tbc_cr['fields']['distrito']
                    sheet2.cell(row=cont, column=4).value = tbc_cr['fields']['documento']
                    sheet2.cell(row=cont, column=5).value = tbc_cr['fields']['financiador']
                    sheet2.cell(row=cont, column=6).value = tbc_cr['fields']['eess_v1']
                    sheet2.cell(row=cont, column=6).alignment = Alignment(wrap_text=True)
                    sheet2.cell(row=cont, column=7).value = tbc_cr['fields']['visita1']
                    sheet2.cell(row=cont, column=8).value = tbc_cr['fields']['eess_v2']
                    sheet2.cell(row=cont, column=8).alignment = Alignment(wrap_text=True)
                    sheet2.cell(row=cont, column=9).value = tbc_cr['fields']['visita2']
                    sheet2.cell(row=cont, column=10).value = tbc_cr['fields']['diasv']
                    cont = cont+1
                    num = num+1


            sheet3.row_dimensions[2].height = 36
            sheet3.column_dimensions['A'].width = 7
            sheet3.column_dimensions['B'].width = 24
            sheet3.column_dimensions['C'].width = 40
            sheet3.column_dimensions['D'].width = 13
            sheet3.column_dimensions['E'].width = 10
            sheet3.column_dimensions['F'].width = 10
            sheet3.column_dimensions['G'].width = 15

            set_border(self, sheet3, "A2:G2", "medium", "2F75B5")

            sheet3.merge_cells('A2:G2')
            sheet3['A2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            sheet3['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sheet3['A2'] = 'DIRESA PASCO - 0016 TBC-VIH/SIDA - TBC-Familias que reciben consejería para promover prácticas y entornos saludables para contribuir a la disminución de la Tuberculosis y VIH/SIDA (CONSOLIDADO) - ' + request.GET['anio']

            sheet3['A4'] = '#'
            sheet3['A4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['A4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['A4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['A4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['B4'] = 'Provincia'
            sheet3['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['B4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet3['B4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['C4'] = 'Distrito'
            sheet3['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['C4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['C4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['C4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['D4'] = 'Financiador'
            sheet3['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['D4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['D4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['D4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['E4'] = 'Total'
            sheet3['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['E4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['E4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['E4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['F4'] = 'Cumplen'
            sheet3['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['F4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['F4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet3['F4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['G4'] = 'Avance'
            sheet3['G4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['G4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['G4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['G4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            cont = 5
            cant = len(dataCont)
            num=1
            if cant > 0:
                for tbc_c in dataCont:
                    sheet3.cell(row=cont, column=1).value = num
                    sheet3.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    sheet3.cell(row=cont, column=2).value = tbc_c['fields']['provincia']
                    sheet3.cell(row=cont, column=3).value = tbc_c['fields']['distrito']
                    sheet3.cell(row=cont, column=4).value = tbc_c['fields']['financiador']
                    sheet3.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
                    sheet3.cell(row=cont, column=5).value = tbc_c['fields']['den']
                    sheet3.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    sheet3.cell(row=cont, column=6).value = tbc_c['fields']['num']
                    sheet3.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    avance = round((tbc_c['fields']['num']/tbc_c['fields']['den'])*100, 1) if (tbc_c['fields']['den'] != 0) and (tbc_c['fields']['den'] is not None) else 0
                    sheet3.cell(row=cont, column=7).value = str(avance) + ' %'
                    sheet3.cell(row=cont, column=7).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO TBC - FAMILIAS QUE RECIB. CONSEJ. PARA DISMINUIR TBC Y VIH.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            sheet2.title = 'CRUCE'
            sheet3.title = 'CONSOLIDADO'
            wb.save(response)
            return response

        elif request.GET['tipo'] == 'tbcvih_vih':
            sheet2 = wb.create_sheet('CRUCE')
            sheet3 = wb.create_sheet('CONTEO')

            if request.GET['prov'] == 'TODOS':
                dataNom = tbcvih016_vih_n.objects.filter(anio=request.GET['anio']).order_by('provincia')
                dataCruce = tbcvih016_vih_cr.objects.filter(anio=request.GET['anio']).order_by('provincia')
                dataCont = tbcvih016_vih_c.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = tbcvih016_vih_n.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
                dataCruce = tbcvih016_vih_cr.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
                dataCont = tbcvih016_vih_c.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = tbcvih016_vih_n.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCruce = tbcvih016_vih_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCont = tbcvih016_vih_c.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = tbcvih016_vih_n.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
                dataCruce = tbcvih016_vih_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCont = tbcvih016_vih_c.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
            dataCruce = json.loads(serializers.serialize('json', dataCruce, indent=2, use_natural_foreign_keys=True))
            dataCont = json.loads(serializers.serialize('json', dataCont, indent=2, use_natural_foreign_keys=True))

            set_border(self, ws, "A2:K2", "medium", "2F75B5")
            set_border(self, ws, "A3:K3", "medium", "366092")
            set_border(self, ws, "A5:K5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 36
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 8
            ws.column_dimensions['J'].width = 50
            ws.column_dimensions['K'].width = 12

            ws.merge_cells('B2:K2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: 0016 TBC-VIH/SIDA - VIH - Familias que reciben consejería para promover prácticas y entornos saludables para contribuir a la disminución de la Tuberculosis y VIH/SIDA - ' + request.GET['anio']

            ws.merge_cells('B3:K3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:K5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Documento'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Mes'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Financiador'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Fecha Aten.'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Visita'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['J7'] = 'Sub Producto'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['K7'] = 'Reg'
            ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for vih in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = vih['fields']['provincia']
                    ws.cell(row=cont, column=3).value = vih['fields']['distrito']
                    ws.cell(row=cont, column=4).value = vih['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = vih['fields']['documento']
                    ws.cell(row=cont, column=6).value = vih['fields']['mes']
                    ws.cell(row=cont, column=7).value = vih['fields']['financiador']
                    ws.cell(row=cont, column=8).value = vih['fields']['fec_atencion']
                    ws.cell(row=cont, column=9).value = vih['fields']['visita']
                    ws.cell(row=cont, column=10).value = vih['fields']['subproduct']
                    ws.cell(row=cont, column=10).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=11).value = vih['fields']['reg']

                    cont = cont+1
                    num = num+1

            sheet2.row_dimensions[2].height = 36
            sheet2.column_dimensions['A'].width = 7
            sheet2.column_dimensions['B'].width = 24
            sheet2.column_dimensions['C'].width = 24
            sheet2.column_dimensions['D'].width = 10
            sheet2.column_dimensions['E'].width = 15
            sheet2.column_dimensions['F'].width = 26
            sheet2.column_dimensions['G'].width = 12
            sheet2.column_dimensions['H'].width = 26
            sheet2.column_dimensions['I'].width = 12

            set_border(self, sheet2, "A2:I2", "medium", "2F75B5")

            sheet2.merge_cells('A2:I2')
            sheet2['A2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            sheet2['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sheet2['A2'] = 'DIRESA PASCO - DEIT: 0016 TBC-VIH/SIDA - VIH - Familias que reciben consejería para promover prácticas y entornos saludables para contribuir a la disminución de la Tuberculosis y VIH/SIDA (CRUCE) - ' + request.GET['anio']

            sheet2['A4'] = '#'
            sheet2['A4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['A4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['A4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['A4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['B4'] = 'Provincia'
            sheet2['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['B4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['B4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['C4'] = 'Distrito'
            sheet2['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['C4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['C4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['C4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['D4'] = 'Documento'
            sheet2['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['D4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['D4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['D4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['E4'] = 'Financiador'
            sheet2['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['E4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['E4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['E4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['F4'] = 'EESS Vista1'
            sheet2['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['F4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['F4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['F4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['G4'] = 'Visita1'
            sheet2['G4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['G4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['G4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['G4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['H4'] = 'EESS Vista2'
            sheet2['H4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['H4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['H4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['H4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['I4'] = 'Visita2'
            sheet2['I4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['I4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['I4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['I4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 5
            cant = len(dataCruce)
            num=1
            if cant > 0:
                for vih_cr in dataCruce:
                    sheet2.cell(row=cont, column=1).value = num
                    sheet2.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=2).value = vih_cr['fields']['provincia']
                    sheet2.cell(row=cont, column=3).value = vih_cr['fields']['distrito']
                    sheet2.cell(row=cont, column=4).value = vih_cr['fields']['documento']
                    sheet2.cell(row=cont, column=5).value = vih_cr['fields']['financiador']
                    sheet2.cell(row=cont, column=6).value = vih_cr['fields']['eess_v1']
                    sheet2.cell(row=cont, column=6).alignment = Alignment(wrap_text=True)
                    sheet2.cell(row=cont, column=7).value = vih_cr['fields']['visita1']
                    sheet2.cell(row=cont, column=8).value = vih_cr['fields']['eess_v2']
                    sheet2.cell(row=cont, column=8).alignment = Alignment(wrap_text=True)
                    sheet2.cell(row=cont, column=9).value = vih_cr['fields']['visita2']

                    cont = cont+1
                    num = num+1


            sheet3.row_dimensions[2].height = 36
            sheet3.column_dimensions['A'].width = 7
            sheet3.column_dimensions['B'].width = 24
            sheet3.column_dimensions['C'].width = 40
            sheet3.column_dimensions['D'].width = 13
            sheet3.column_dimensions['E'].width = 10
            sheet3.column_dimensions['F'].width = 10
            sheet3.column_dimensions['G'].width = 15

            set_border(self, sheet3, "A2:G2", "medium", "2F75B5")

            sheet3.merge_cells('A2:G2')
            sheet3['A2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            sheet3['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sheet3['A2'] = 'DIRESA PASCO - 0016 TBC-VIH/SIDA - VIH - Familias que reciben consejería para promover prácticas y entornos saludables para contribuir a la disminución de la Tuberculosis y VIH/SIDA (CONSOLIDADO) - ' + request.GET['anio']

            sheet3['A4'] = '#'
            sheet3['A4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['A4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['A4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['A4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['B4'] = 'Provincia'
            sheet3['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['B4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet3['B4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['C4'] = 'Distrito'
            sheet3['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['C4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['C4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['C4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['D4'] = 'Financiador'
            sheet3['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['D4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['D4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['D4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['E4'] = 'Total'
            sheet3['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['E4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['E4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['E4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['F4'] = 'Cumplen'
            sheet3['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['F4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['F4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet3['F4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet3['G4'] = 'Avance'
            sheet3['G4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet3['G4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet3['G4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet3['G4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            cont = 5
            cant = len(dataCont)
            num=1
            if cant > 0:
                for vih_c in dataCont:
                    sheet3.cell(row=cont, column=1).value = num
                    sheet3.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    sheet3.cell(row=cont, column=2).value = vih_c['fields']['provincia']
                    sheet3.cell(row=cont, column=3).value = vih_c['fields']['distrito']
                    sheet3.cell(row=cont, column=4).value = vih_c['fields']['financiador']
                    sheet3.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
                    sheet3.cell(row=cont, column=5).value = vih_c['fields']['den']
                    sheet3.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    sheet3.cell(row=cont, column=6).value = vih_c['fields']['num']
                    sheet3.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    avance = round((vih_c['fields']['num']/vih_c['fields']['den'])*100, 1) if (vih_c['fields']['den'] != 0) and (vih_c['fields']['den'] is not None) else 0
                    sheet3.cell(row=cont, column=7).value = str(avance) + ' %'
                    sheet3.cell(row=cont, column=7).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO VIH - FAMILIAS QUE RECIB. CONSEJ. PARA DISMINUIR TBC Y VIH.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            sheet2.title = 'CRUCE'
            sheet3.title = 'CONSOLIDADO'
            wb.save(response)
            return response

        elif request.GET['tipo'] == 'met017':
            sheet2 = wb.create_sheet('CRUCE')

            if request.GET['prov'] == 'TODOS':
                dataNom = met017_met_n.objects.filter(anio=request.GET['anio']).order_by('provincia')
                dataCruce = met017_met_cr.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = met017_met_n.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
                dataCruce = met017_met_cr.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = met017_met_n.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
                dataCruce = met017_met_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = met017_met_n.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
                dataCruce = met017_met_cr.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')

            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
            dataCruce = json.loads(serializers.serialize('json', dataCruce, indent=2, use_natural_foreign_keys=True))

            set_border(self, ws, "A2:K2", "medium", "2F75B5")
            set_border(self, ws, "A3:K3", "medium", "366092")
            set_border(self, ws, "A5:K5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 50
            ws.column_dimensions['K'].width = 5

            ws.merge_cells('B2:K2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: 0017 METAXÈNICAS Y ZOONOTICAS Familias que reciben sesiones demostrativas desarrollan prácticas saludables para la prevención de las enfermedades metaxénicas. - ' + request.GET['anio']

            ws.merge_cells('B3:K3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:K5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Documento'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Mes'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Fecha Aten.'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Motivo'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Sesión'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['J7'] = 'Sub Producto'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['K7'] = 'Reg'
            ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for metax in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = metax['fields']['provincia']
                    ws.cell(row=cont, column=3).value = metax['fields']['distrito']
                    ws.cell(row=cont, column=4).value = metax['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = metax['fields']['documento']
                    ws.cell(row=cont, column=6).value = metax['fields']['mes']
                    ws.cell(row=cont, column=7).value = metax['fields']['fec_atencion']
                    ws.cell(row=cont, column=8).value = metax['fields']['motivo']
                    ws.cell(row=cont, column=9).value = metax['fields']['sesion']
                    ws.cell(row=cont, column=10).value = metax['fields']['subproduct']
                    ws.cell(row=cont, column=10).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=11).value = metax['fields']['reg']

                    cont = cont+1
                    num = num+1

            sheet2.row_dimensions[2].height = 23
            sheet2.column_dimensions['A'].width = 7
            sheet2.column_dimensions['B'].width = 24
            sheet2.column_dimensions['C'].width = 24
            sheet2.column_dimensions['D'].width = 10
            sheet2.column_dimensions['E'].width = 24
            sheet2.column_dimensions['F'].width = 5
            sheet2.column_dimensions['G'].width = 12
            sheet2.column_dimensions['H'].width = 12
            sheet2.column_dimensions['I'].width = 12
            sheet2.column_dimensions['J'].width = 12
            sheet2.column_dimensions['K'].width = 12
            sheet2.column_dimensions['L'].width = 12
            sheet2.column_dimensions['M'].width = 12
            sheet2.column_dimensions['N'].width = 12
            sheet2.column_dimensions['O'].width = 12
            sheet2.column_dimensions['P'].width = 12
            sheet2.column_dimensions['Q'].width = 12
            sheet2.column_dimensions['R'].width = 12
            sheet2.column_dimensions['S'].width = 12
            sheet2.column_dimensions['T'].width = 12
            sheet2.column_dimensions['U'].width = 12
            sheet2.column_dimensions['V'].width = 12
            sheet2.column_dimensions['W'].width = 12
            sheet2.column_dimensions['X'].width = 12
            sheet2.column_dimensions['Y'].width = 12
            sheet2.column_dimensions['Z'].width = 12

            set_border(self, sheet2, "A2:Z2", "medium", "2F75B5")

            sheet2.merge_cells('A2:Z2')
            sheet2['A2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            sheet2['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sheet2['A2'] = 'DIRESA PASCO - DEIT: 0017 METAXÈNICAS Y ZOONOTICAS Familias que reciben sesiones demostrativas desarrollan prácticas saludables para la prevención de las enfermedades metaxénicas. (CRUCE) - ' + request.GET['anio']

            sheet2['A4'] = '#'
            sheet2['A4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['A4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['A4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['A4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['B4'] = 'Provincia'
            sheet2['B4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['B4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['B4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['C4'] = 'Distrito'
            sheet2['C4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['C4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['C4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['C4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['D4'] = 'Documento'
            sheet2['D4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['D4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['D4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['D4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['E4'] = 'Sub Producto'
            sheet2['E4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['E4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['E4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet2['E4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['F4'] = 'Reg'
            sheet2['F4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['F4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['F4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['F4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['G4'] = 'Dengue1'
            sheet2['G4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['G4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['G4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['G4'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            sheet2['H4'] = 'Dengue2'
            sheet2['H4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['H4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['H4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['H4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['I4'] = 'Chikin1'
            sheet2['I4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['I4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['I4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['I4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['J4'] = 'Chikin2'
            sheet2['J4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['J4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['J4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['J4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['K4'] = 'Zoorabia1'
            sheet2['K4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['K4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['K4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['K4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['L4'] = 'Zoorabia2'
            sheet2['L4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['L4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['L4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['L4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['M4'] = 'Equino1'
            sheet2['M4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['M4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['M4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['M4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['N4'] = 'Equino2'
            sheet2['N4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['N4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['N4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['N4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['O4'] = 'F Amar1'
            sheet2['O4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['O4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['O4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['O4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['P4'] = 'F Amar2'
            sheet2['P4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['P4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['P4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['P4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['Q4'] = 'Leishma1'
            sheet2['Q4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['Q4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['Q4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['Q4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['R4'] = 'Leishma2'
            sheet2['R4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['R4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['R4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['R4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['S4'] = 'Malaria1'
            sheet2['S4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['S4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['S4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['S4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['T4'] = 'Malaria2'
            sheet2['T4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['T4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['T4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['T4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['U4'] = 'Peste1'
            sheet2['U4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['U4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['U4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['U4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['V4'] = 'Peste2'
            sheet2['V4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['V4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['V4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['V4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['W4'] = 'Tifus1'
            sheet2['W4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['W4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['W4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['W4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['X4'] = 'Tifus2'
            sheet2['X4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['X4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['X4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['X4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['Y4'] = 'Zika1'
            sheet2['Y4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['Y4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['Y4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['Y4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            sheet2['Z4'] = 'Zika2'
            sheet2['Z4'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            sheet2['Z4'].font = Font(name='Aptos Narrow', size=10, bold=True)
            sheet2['Z4'].alignment = Alignment(horizontal="center", vertical="center")
            sheet2['Z4'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            cont = 5
            cant = len(dataCruce)
            num=1
            if cant > 0:
                for vih_cr in dataCruce:
                    sheet2.cell(row=cont, column=1).value = num
                    sheet2.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    sheet2.cell(row=cont, column=2).value = vih_cr['fields']['provincia']
                    sheet2.cell(row=cont, column=3).value = vih_cr['fields']['distrito']
                    sheet2.cell(row=cont, column=4).value = vih_cr['fields']['documento']
                    sheet2.cell(row=cont, column=5).value = vih_cr['fields']['subproduct']
                    sheet2.cell(row=cont, column=5).alignment = Alignment(wrap_text=True)
                    sheet2.cell(row=cont, column=6).value = vih_cr['fields']['reg']
                    sheet2.cell(row=cont, column=7).value = vih_cr['fields']['dengue1']
                    sheet2.cell(row=cont, column=8).value = vih_cr['fields']['dengue2']
                    sheet2.cell(row=cont, column=9).value = vih_cr['fields']['chikin1']
                    sheet2.cell(row=cont, column=10).value = vih_cr['fields']['chikin2']
                    sheet2.cell(row=cont, column=11).value = vih_cr['fields']['zoorabia1']
                    sheet2.cell(row=cont, column=12).value = vih_cr['fields']['zoorabia2']
                    sheet2.cell(row=cont, column=13).value = vih_cr['fields']['equino1']
                    sheet2.cell(row=cont, column=14).value = vih_cr['fields']['equino2']
                    sheet2.cell(row=cont, column=15).value = vih_cr['fields']['f_amar1']
                    sheet2.cell(row=cont, column=16).value = vih_cr['fields']['f_amar2']
                    sheet2.cell(row=cont, column=17).value = vih_cr['fields']['leishma1']
                    sheet2.cell(row=cont, column=18).value = vih_cr['fields']['chikin1']
                    sheet2.cell(row=cont, column=19).value = vih_cr['fields']['leishma2']
                    sheet2.cell(row=cont, column=20).value = vih_cr['fields']['malaria1']
                    sheet2.cell(row=cont, column=21).value = vih_cr['fields']['malaria2']
                    sheet2.cell(row=cont, column=22).value = vih_cr['fields']['peste1']
                    sheet2.cell(row=cont, column=23).value = vih_cr['fields']['peste2']
                    sheet2.cell(row=cont, column=24).value = vih_cr['fields']['tifus1']
                    sheet2.cell(row=cont, column=25).value = vih_cr['fields']['tifus2']
                    sheet2.cell(row=cont, column=26).value = vih_cr['fields']['zika1']
                    sheet2.cell(row=cont, column=27).value = vih_cr['fields']['zika2']

                    cont = cont+1
                    num = num+1


            nombre_archivo = "DEIT_PASCO METAXÈNICAS Y ZOONOTICAS FAMILIAS QUE RECIB. SESIONES DEMOST.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            sheet2.title = 'CRUCE'
            wb.save(response)
            return response

        elif request.GET['tipo'] == 'cc':
            set_border(self, ws, "A2:k2", "medium", "2F75B5")
            set_border(self, ws, "A3:k3", "medium", "366092")
            set_border(self, ws, "A5:k5", "medium", "D9D9D9")

            ws.row_dimensions[2].height = 23
            ws.column_dimensions['A'].width = 7
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 24
            ws.column_dimensions['E'].width = 5
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 45
            ws.column_dimensions['K'].width = 45

            ws.merge_cells('B2:k2')
            ws['B2'].font = Font(name='Aptos Narrow', size=12, bold=True, color='2F75B5')
            ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws['B2'] = 'DIRESA PASCO - DEIT: PROMOCIÓN DE LA SALUD (CONTROL DE CALIDAD)  - ' + request.GET['anio']

            ws.merge_cells('B3:k3')
            ws['B3'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
            ws['B3'] = 'CODIFICACION: '

            ws.merge_cells('A5:k5')
            ws['A5'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
            ws['A5'] = 'Fuente: BD_HISMINSA con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

            ws['A7'] = '#'
            ws['A7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['A7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['B7'] = 'Provincia'
            ws['B7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['B7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['B7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['B7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['C7'] = 'Distrito'
            ws['C7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['C7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['C7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['D7'] = 'Establecimiento'
            ws['D7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['D7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['D7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['E7'] = 'Mes'
            ws['E7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['E7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['F7'] = 'Documento'
            ws['F7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['F7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['F7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['F7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['G7'] = 'Fecha Nacido'
            ws['G7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['G7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['G7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['H7'] = 'Fecha Aten.'
            ws['H7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['H7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['H7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['H7'].fill = PatternFill(start_color='CBD5F5', end_color='CBD5F5', fill_type='solid')

            ws['I7'] = 'Reg. Manual'
            ws['I7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['I7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['I7'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws['I7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['J7'] = 'Sub Producto'
            ws['J7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['J7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['J7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            ws['K7'] = 'Observación'
            ws['K7'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
            ws['K7'].font = Font(name='Aptos Narrow', size=10, bold=True)
            ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].fill = PatternFill(start_color='abc7fb', end_color='abc7fb', fill_type='solid')

            if request.GET['prov'] == 'TODOS':
                dataNom = cc.objects.filter(anio=request.GET['anio']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] == 'TODOS':
                dataNom = cc.objects.filter(anio=request.GET['anio'], cod_prov=request.GET['prov']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] == 'TODOS':
                dataNom = cc.objects.filter(anio=request.GET['anio'], cod_dist=request.GET['dist']).order_by('provincia')
            elif request.GET['prov'] != 'TODOS' and request.GET['dist'] != 'TODOS' and request.GET['eess'] != 'TODOS':
                dataNom = cc.objects.filter(anio=request.GET['anio'], cod_eess=request.GET['eess']).order_by('provincia')
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            cont = 8
            cant = len(dataNom)
            num=1
            if cant > 0:
                for cc_nom in dataNom:
                    ws.cell(row=cont, column=1).value = num
                    ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=2).value = cc_nom['fields']['provincia']
                    ws.cell(row=cont, column=3).value = cc_nom['fields']['distrito']
                    ws.cell(row=cont, column=4).value = cc_nom['fields']['eess']
                    ws.cell(row=cont, column=4).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=5).value = cc_nom['fields']['mes']
                    ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=6).value = cc_nom['fields']['documento']
                    ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=7).value = cc_nom['fields']['fec_nac']
                    ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=8).value = cc_nom['fields']['fec_atencion']
                    ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=9).value = cc_nom['fields']['reg_manual']
                    ws.cell(row=cont, column=9).alignment = Alignment(horizontal="center")
                    ws.cell(row=cont, column=10).value = cc_nom['fields']['subproduct']
                    ws.cell(row=cont, column=10).alignment = Alignment(wrap_text=True)
                    ws.cell(row=cont, column=11).value = cc_nom['fields']['observacion']
                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal="center")

                    cont = cont+1
                    num = num+1

            nombre_archivo = "DEIT_PASCO PROMSA CONTROL DE CALIDAD.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            ws.title = 'NOMINAL'
            wb.save(response)
            return response
